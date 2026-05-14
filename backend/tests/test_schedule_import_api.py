import os
import tempfile
import unittest
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.dependencies import get_db, require_school_admin
from app.main import app
from app.models.class_ import Class
from app.models.schedule_draft import ScheduleDraft
from app.models.schedule_draft_item import ScheduleDraftItem
from app.models.schedule_period import SchedulePeriod
from app.models.school import School
from app.models.subject import Subject
from app.models.teacher_class_subject import TeacherClassSubject
from app.models.user import User


class ScheduleImportApiTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
        self.tmp.close()
        self.engine = create_engine(f"sqlite:///{self.tmp.name}", connect_args={"check_same_thread": False})
        self.SessionTesting = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        Base.metadata.create_all(bind=self.engine)

        self.db = self.SessionTesting()
        self.school = School(name="导入测试学校", location="测试", school_level="middle")
        self.db.add(self.school)
        self.db.commit()
        self.db.refresh(self.school)

        self.admin = User(username="schedule_import_admin", password_hash="x", role="school_admin", school_id=self.school.id)
        self.teacher = User(username="schedule_import_teacher", password_hash="x", role="teacher", school_id=self.school.id)
        self.db.add_all([self.admin, self.teacher])
        self.db.commit()
        self.db.refresh(self.admin)
        self.db.refresh(self.teacher)

        self.class_obj = Class(name="1班", grade="八年级", school_id=self.school.id)
        self.subject = Subject(name="数学", code="MATH", grades="八年级", school_id=self.school.id)
        self.period = SchedulePeriod(
            name="第1节",
            start_time="08:00",
            end_time="08:45",
            school_id=self.school.id,
            sort_order=1,
            is_active=True,
            include_in_auto_schedule=True,
        )
        self.db.add_all([self.class_obj, self.subject, self.period])
        self.db.commit()
        self.db.refresh(self.class_obj)
        self.db.refresh(self.subject)
        self.db.refresh(self.period)
        self.db.add(
            TeacherClassSubject(
                school_id=self.school.id,
                class_id=self.class_obj.id,
                subject_id=self.subject.id,
                teacher_id=self.teacher.id,
            )
        )
        self.db.commit()

        def override_get_db():
            db = self.SessionTesting()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[require_school_admin] = lambda: self.admin
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()
        os.unlink(self.tmp.name)

    def test_get_missing_schedule_import_returns_business_404(self):
        response = self.client.get("/api/schedule/imports/999")

        self.assertEqual(response.status_code, 404)
        self.assertEqual(response.json()["detail"], "import not found")

    def test_excel_import_matches_teacher_from_arrangement(self):
        data = self._create_excel_import()

        self.assertEqual(data["status"], "needs_review")
        self.assertEqual(data["summary"]["total_slots"], 1)
        self.assertEqual(data["summary"]["recognized_slots"], 1)
        self.assertEqual(data["summary"]["teacher_unmatched_slots"], 0)

        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        self.assertEqual(items_response.status_code, 200)
        items = items_response.json()["items"]
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["subject_id"], self.subject.id)
        self.assertEqual(items[0]["teacher_id"], self.teacher.id)
        self.assertEqual(items[0]["teacher_match_status"], "matched")
        self.assertEqual(items[0]["teacher_match_source"], "teaching_arrangement")

    def test_excel_import_matches_common_subject_abbreviation(self):
        data = self._create_excel_import(subject_text="数")

        self.assertEqual(data["summary"]["recognized_slots"], 1)
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        item = items_response.json()["items"][0]
        self.assertEqual(item["subject_id"], self.subject.id)
        self.assertEqual(item["subject_name"], "数学")

    def test_grade_scope_import_reads_each_sheet_as_matching_class(self):
        class_one = Class(name="七一班", grade="七年级", school_id=self.school.id)
        class_two = Class(name="七二班", grade="七年级", school_id=self.school.id)
        self.db.add_all([class_one, class_two])
        self.db.commit()
        self.db.refresh(class_one)
        self.db.refresh(class_two)
        self.db.add_all(
            [
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_one.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_two.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
            ]
        )
        self.db.commit()

        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "七1班"
        first_sheet.append(["节次", "周一"])
        first_sheet.append(["第1节", "数学"])
        second_sheet = workbook.create_sheet("七二班")
        second_sheet.append(["节次", "周二"])
        second_sheet.append(["第1节", "数学"])
        ignored_sheet = workbook.create_sheet("无关说明")
        ignored_sheet.append(["节次", "周一"])
        ignored_sheet.append(["第1节", "数学"])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "七年级", "scope": "grade"},
            files={
                "file": (
                    "七年级总课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["scope"], "grade")
        self.assertEqual(data["summary"]["total_slots"], 2)
        self.assertIn("无关说明", data["message"])

        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        items = items_response.json()["items"]
        self.assertEqual({item["class_id"] for item in items}, {class_one.id, class_two.id})
        self.assertEqual([item["weekday"] for item in items], [1, 2])

    def test_grade_scope_import_matches_sheet_without_class_suffix_and_finds_header_row(self):
        class_obj = Class(name="八一班", grade="八年级", school_id=self.school.id)
        self.db.add(class_obj)
        self.db.commit()
        self.db.refresh(class_obj)
        self.db.add(
            TeacherClassSubject(
                school_id=self.school.id,
                class_id=class_obj.id,
                subject_id=self.subject.id,
                teacher_id=self.teacher.id,
            )
        )
        self.db.commit()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "八一"
        sheet.append(["课 程 表", None, None, None, None, None])
        sheet.append(["八一", None, None, None, None, None])
        sheet.append(["节次 星期", "星期一", "星期二", "星期三", "星期四", "星期五"])
        sheet.append(["第1节", "数学", None, None, None, None])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": "grade"},
            files={
                "file": (
                    "八年级总课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["summary"]["total_slots"], 1)
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        item = items_response.json()["items"][0]
        self.assertEqual(item["class_id"], class_obj.id)
        self.assertEqual(item["weekday"], 1)

    def test_excel_import_maps_template_periods_by_active_lesson_order(self):
        self_study = SchedulePeriod(
            name="早自习",
            start_time="07:00",
            end_time="07:40",
            school_id=self.school.id,
            sort_order=1,
            is_active=True,
            include_in_auto_schedule=False,
        )
        self.db.add(self_study)
        self.period.name = "第一节"
        self.period.sort_order = 2
        self.db.commit()

        data = self._create_excel_import()

        self.assertEqual(data["summary"]["total_slots"], 1)
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        item = items_response.json()["items"][0]
        self.assertEqual(item["period_id"], self.period.id)

    def test_excel_import_uses_legacy_global_periods_when_school_has_none(self):
        self.period.school_id = None
        self.db.commit()

        data = self._create_excel_import()

        self.assertEqual(data["summary"]["total_slots"], 1)
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        item = items_response.json()["items"][0]
        self.assertEqual(item["period_id"], self.period.id)

    def test_excel_import_requires_schedulable_periods(self):
        self.period.include_in_auto_schedule = False
        self.db.commit()

        response = self._post_excel_import()

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "请先配置参与排课的节次")

    def test_excel_import_reads_merged_time_band_layout(self):
        file_teacher = User(username="张老师", password_hash="x", role="teacher", school_id=self.school.id)
        self.db.add(file_teacher)
        self.db.commit()
        self.db.refresh(file_teacher)

        workbook = Workbook()
        sheet = workbook.active
        sheet.merge_cells("A1:C1")
        sheet["A1"] = "课 程 星期 时间"
        sheet["D1"] = "星期一"
        sheet["E1"] = "星期二"
        sheet.merge_cells("A2:A4")
        sheet["A2"] = "上午"
        sheet.merge_cells("B2:C2")
        sheet["B2"] = "1"
        sheet["D2"] = "数学\n(张老师)"
        sheet["E2"] = "数学"
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": "class", "class_id": str(self.class_obj.id)},
            files={
                "file": (
                    "合并单元格课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["summary"]["total_slots"], 2)

        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        items = items_response.json()["items"]
        self.assertEqual([item["weekday"] for item in items], [1, 2])
        self.assertEqual([item["period_id"] for item in items], [self.period.id, self.period.id])
        self.assertEqual(items[0]["teacher_id"], file_teacher.id)
        self.assertEqual(items[0]["teacher_match_source"], "file_recognition")

    def test_download_excel_template(self):
        response = self.client.get("/api/schedule/imports/template")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["content-type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment;", response.headers["content-disposition"])

    def test_patch_item_and_create_draft(self):
        data = self._create_excel_import()
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        item_id = items_response.json()["items"][0]["id"]

        patch_response = self.client.patch(
            f"/api/schedule/imports/{data['id']}/items/{item_id}",
            json={"teacher_id": self.teacher.id},
        )
        self.assertEqual(patch_response.status_code, 200)
        self.assertEqual(patch_response.json()["teacher_match_source"], "manual")

        draft_response = self.client.post(f"/api/schedule/imports/{data['id']}/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_id = draft_response.json()["draft_id"]
        self.assertEqual(draft_response.json()["status"], "draft_created")

        db = self.SessionTesting()
        try:
            draft = db.query(ScheduleDraft).filter(ScheduleDraft.id == draft_id).first()
            self.assertIsNotNone(draft)
            self.assertEqual(draft.status, "draft")
            draft_item = db.query(ScheduleDraftItem).filter(ScheduleDraftItem.draft_id == draft_id).first()
            self.assertIsNotNone(draft_item)
            self.assertEqual(draft_item.subject_id, self.subject.id)
            self.assertEqual(draft_item.teacher_id, self.teacher.id)
        finally:
            db.close()

    def test_patch_item_persists_teacher_class_subject_binding_for_next_import(self):
        self.db.query(TeacherClassSubject).delete()
        self.db.commit()

        first_import = self._create_excel_import()
        first_items_response = self.client.get(f"/api/schedule/imports/{first_import['id']}/items")
        first_item = first_items_response.json()["items"][0]
        self.assertEqual(first_item["teacher_id"], None)
        self.assertIn("teacher_unmatched", first_item["issue_flags"])

        patch_response = self.client.patch(
            f"/api/schedule/imports/{first_import['id']}/items/{first_item['id']}",
            json={"teacher_id": self.teacher.id},
        )
        self.assertEqual(patch_response.status_code, 200)
        duplicate_patch_response = self.client.patch(
            f"/api/schedule/imports/{first_import['id']}/items/{first_item['id']}",
            json={"teacher_id": self.teacher.id},
        )
        self.assertEqual(duplicate_patch_response.status_code, 200)

        db = self.SessionTesting()
        try:
            bindings = (
                db.query(TeacherClassSubject)
                .filter(
                    TeacherClassSubject.school_id == self.school.id,
                    TeacherClassSubject.class_id == self.class_obj.id,
                    TeacherClassSubject.subject_id == self.subject.id,
                    TeacherClassSubject.teacher_id == self.teacher.id,
                )
                .all()
            )
            self.assertEqual(len(bindings), 1)
        finally:
            db.close()

        second_import = self._create_excel_import()
        second_items_response = self.client.get(f"/api/schedule/imports/{second_import['id']}/items")
        second_item = second_items_response.json()["items"][0]
        self.assertEqual(second_item["teacher_id"], self.teacher.id)
        self.assertEqual(second_item["teacher_match_source"], "teaching_arrangement")

    def test_create_draft_rejects_teacher_time_conflicts_with_business_error(self):
        class_one = Class(name="八一班", grade="八年级", school_id=self.school.id)
        class_two = Class(name="八二班", grade="八年级", school_id=self.school.id)
        self.db.add_all([class_one, class_two])
        self.db.commit()
        self.db.refresh(class_one)
        self.db.refresh(class_two)
        self.db.add_all(
            [
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_one.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_two.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
            ]
        )
        self.db.commit()

        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "八一"
        first_sheet.append(["节次", "周一"])
        first_sheet.append(["第1节", "数学"])
        second_sheet = workbook.create_sheet("八二")
        second_sheet.append(["节次", "周一"])
        second_sheet.append(["第1节", "数学"])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        import_response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": "grade"},
            files={
                "file": (
                    "八年级总课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        self.assertEqual(import_response.status_code, 201)

        draft_response = self.client.post(f"/api/schedule/imports/{import_response.json()['id']}/draft")

        self.assertEqual(draft_response.status_code, 400)
        self.assertIn("教师同一时间多班冲突", draft_response.json()["detail"])

    def test_import_marks_teacher_time_conflicts_and_clears_after_correction(self):
        class_one = Class(name="八一班", grade="八年级", school_id=self.school.id)
        class_two = Class(name="八二班", grade="八年级", school_id=self.school.id)
        teacher_two = User(username="schedule_import_teacher_two", password_hash="x", role="teacher", school_id=self.school.id)
        self.db.add_all([class_one, class_two, teacher_two])
        self.db.commit()
        self.db.refresh(class_one)
        self.db.refresh(class_two)
        self.db.refresh(teacher_two)
        self.db.add_all(
            [
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_one.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
                TeacherClassSubject(
                    school_id=self.school.id,
                    class_id=class_two.id,
                    subject_id=self.subject.id,
                    teacher_id=self.teacher.id,
                ),
            ]
        )
        self.db.commit()

        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "八一"
        first_sheet.append(["节次", "周一"])
        first_sheet.append(["第1节", "数学"])
        second_sheet = workbook.create_sheet("八二")
        second_sheet.append(["节次", "周一"])
        second_sheet.append(["第1节", "数学"])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        import_response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": "grade"},
            files={
                "file": (
                    "八年级总课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(import_response.status_code, 201)
        data = import_response.json()
        self.assertEqual(data["summary"]["teacher_time_conflict_slots"], 2)
        items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        items = items_response.json()["items"]
        self.assertEqual(len(items), 2)
        for item in items:
            self.assertIn("teacher_time_conflict", item["issue_flags"])
            self.assertEqual(len(item["conflict_items"]), 1)
            self.assertNotEqual(item["class_id"], item["conflict_items"][0]["class_id"])

        patch_response = self.client.patch(
            f"/api/schedule/imports/{data['id']}/items/{items[0]['id']}",
            json={"teacher_id": teacher_two.id},
        )
        self.assertEqual(patch_response.status_code, 200)
        self.assertNotIn("teacher_time_conflict", patch_response.json()["issue_flags"])

        updated_items_response = self.client.get(f"/api/schedule/imports/{data['id']}/items")
        updated_items = updated_items_response.json()["items"]
        self.assertEqual([item["issue_flags"] for item in updated_items], [[], []])
        updated_import_response = self.client.get(f"/api/schedule/imports/{data['id']}")
        self.assertEqual(updated_import_response.json()["summary"]["teacher_time_conflict_slots"], 0)

    def test_image_upload_is_rejected_because_only_excel_is_supported(self):
        response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": "class", "class_id": str(self.class_obj.id)},
            files={"file": ("schedule.png", b"not really an image", "image/png")},
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["detail"], "仅支持 Excel 课表导入")

    def _create_excel_import(self, *, subject_text: str = "数学"):
        response = self._post_excel_import(scope="class", subject_text=subject_text)
        self.assertEqual(response.status_code, 201)
        return response.json()

    def _post_excel_import(self, *, scope: str = "class", subject_text: str = "数学"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["节次", "周一"])
        sheet.append(["第1节", subject_text])
        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)

        response = self.client.post(
            "/api/schedule/imports",
            data={"grade": "八年级", "scope": scope, "class_id": str(self.class_obj.id)},
            files={
                "file": (
                    "八年级1班课表.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        return response


if __name__ == "__main__":
    unittest.main()
