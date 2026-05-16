import os
import tempfile
import unittest
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.dependencies import get_db, require_school_admin
from app.main import app
from app.models.class_ import Class
from app.models.lesson_plan import LessonPlan
from app.models.school import School
from app.models.schedule_draft import ScheduleDraft
from app.models.schedule_draft_item import ScheduleDraftItem
from app.models.schedule_period import SchedulePeriod
from app.models.subject import Subject
from app.models.teacher_class_subject import TeacherClassSubject
from app.models.user import User


class SchedulingDraftApiTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
        self.tmp.close()
        self.engine = create_engine(f"sqlite:///{self.tmp.name}", connect_args={"check_same_thread": False})
        self.SessionTesting = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        Base.metadata.create_all(bind=self.engine)

        self.db = self.SessionTesting()
        self.school = School(name='测试学校', location='测试', school_level='middle')
        self.db.add(self.school)
        self.db.commit()
        self.db.refresh(self.school)

        self.admin = User(username='school_admin_draft', password_hash='x', role='school_admin', school_id=self.school.id)
        self.teacher = User(username='teacher_draft', password_hash='x', role='teacher', school_id=self.school.id)
        self.db.add_all([self.admin, self.teacher])
        self.db.commit()
        self.db.refresh(self.admin)
        self.db.refresh(self.teacher)

        self.class_obj = Class(name='1班', grade='八年级', school_id=self.school.id)
        self.subject = Subject(name='数学', code='MATH', grades='八年级', school_id=self.school.id)
        self.db.add_all([self.class_obj, self.subject])
        self.db.commit()
        self.db.refresh(self.class_obj)
        self.db.refresh(self.subject)

        self.db.add_all([
            SchedulePeriod(name='第1节', start_time='08:00', end_time='08:45', school_id=self.school.id, sort_order=1, is_active=True, include_in_auto_schedule=True),
            SchedulePeriod(name='第2节', start_time='08:55', end_time='09:40', school_id=self.school.id, sort_order=2, is_active=True, include_in_auto_schedule=True),
        ])
        self.db.commit()

        self.db.add(TeacherClassSubject(school_id=self.school.id, class_id=self.class_obj.id, subject_id=self.subject.id, teacher_id=self.teacher.id))
        self.db.add(LessonPlan(school_id=self.school.id, grade='八年级', subject_id=self.subject.id, content='{"weekly_hours": 1, "preferred_session": "morning_prefer"}'))
        self.db.commit()

        def override_get_db():
            db = self.SessionTesting()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[require_school_admin] = lambda: self.admin
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()
        os.unlink(self.tmp.name)

    def test_solve_draft_returns_pending_task(self):
        response = self.client.post('/api/schedule/drafts/八年级/solve')
        self.assertEqual(response.status_code, 202)
        self.assertEqual(response.json()['status'], 'pending')

    def test_export_draft_workbook_splits_classes_and_omits_teacher_names(self):
        second_class = Class(name='2班', grade='八年级', school_id=self.school.id)
        second_subject = Subject(name='语文', code='CHN', grades='八年级', school_id=self.school.id)
        self.db.add_all([second_class, second_subject])
        self.db.commit()
        self.db.refresh(second_class)
        self.db.refresh(second_subject)

        draft = ScheduleDraft(school_id=self.school.id, grade='八年级', status='draft', score=100, created_by=self.admin.id)
        self.db.add(draft)
        self.db.flush()
        first_period = self.db.query(SchedulePeriod).filter(SchedulePeriod.sort_order == 1).first()
        second_period = self.db.query(SchedulePeriod).filter(SchedulePeriod.sort_order == 2).first()
        self.db.add_all([
            ScheduleDraftItem(
                draft_id=draft.id,
                class_id=self.class_obj.id,
                teacher_id=self.teacher.id,
                subject_id=self.subject.id,
                weekday=1,
                period_id=first_period.id,
            ),
            ScheduleDraftItem(
                draft_id=draft.id,
                class_id=second_class.id,
                teacher_id=self.teacher.id,
                subject_id=second_subject.id,
                weekday=2,
                period_id=second_period.id,
            ),
        ])
        self.db.commit()

        response = self.client.get(f'/api/schedule/drafts/{draft.id}/export')

        self.assertEqual(response.status_code, 200)
        self.assertIn('schedule-draft-', response.headers['content-disposition'])
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['1班', '2班'])
        first_sheet = workbook['1班']
        second_sheet = workbook['2班']
        self.assertEqual(first_sheet['A1'].value, '课 程 表')
        self.assertEqual(first_sheet['A2'].value, '1班')
        self.assertEqual(first_sheet['A3'].value, '节次')
        self.assertEqual(first_sheet['B4'].value, '数学')
        self.assertTrue(first_sheet['A1'].font.bold)
        self.assertEqual(first_sheet['B4'].alignment.horizontal, 'center')
        self.assertEqual(first_sheet['A3'].border.left.style, 'thin')
        self.assertNotIn('teacher_draft', [cell.value for row in first_sheet.iter_rows() for cell in row])
        self.assertEqual(second_sheet['C5'].value, '语文')

    def test_export_draft_sorts_chinese_class_sheets_by_class_number(self):
        classes = [
            Class(name='八八班', grade='八年级', school_id=self.school.id),
            Class(name='八二班', grade='八年级', school_id=self.school.id),
            Class(name='八一班', grade='八年级', school_id=self.school.id),
        ]
        self.db.add_all(classes)
        self.db.commit()
        for class_item in classes:
            self.db.refresh(class_item)

        draft = ScheduleDraft(school_id=self.school.id, grade='八年级', status='draft', score=100, created_by=self.admin.id)
        self.db.add(draft)
        self.db.flush()
        first_period = self.db.query(SchedulePeriod).filter(SchedulePeriod.sort_order == 1).first()
        self.db.add_all([
            ScheduleDraftItem(
                draft_id=draft.id,
                class_id=class_item.id,
                teacher_id=self.teacher.id,
                subject_id=self.subject.id,
                weekday=index + 1,
                period_id=first_period.id,
            )
            for index, class_item in enumerate(classes)
        ])
        self.db.commit()

        response = self.client.get(f'/api/schedule/drafts/{draft.id}/export')

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ['八一班', '八二班', '八八班'])
