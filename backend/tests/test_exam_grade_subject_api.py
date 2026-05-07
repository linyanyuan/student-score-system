# -*- coding: utf-8 -*-
import io
import os
import tempfile
import unittest

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.dependencies import get_current_user, get_db, require_admin_or_school_admin, require_teacher_or_admin
from app.main import app
from app.models.class_ import Class
from app.models.school import School
from app.models.score import Score
from app.models.subject import Subject
from app.models.student import Student
from app.models.total_rank import TotalRank
from app.models.user import User


class ExamGradeSubjectApiTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".db")
        self.tmp.close()
        self.engine = create_engine(f"sqlite:///{self.tmp.name}", connect_args={"check_same_thread": False})
        self.SessionTesting = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        Base.metadata.create_all(bind=self.engine)

        self.db = self.SessionTesting()
        self.school = School(name="测试学校", location="测试", school_level="middle")
        self.db.add(self.school)
        self.db.commit()
        self.db.refresh(self.school)

        self.admin = User(
            username="school_admin_exam_case",
            password_hash="x",
            role="school_admin",
            school_id=self.school.id,
        )
        self.db.add(self.admin)
        self.db.commit()
        self.db.refresh(self.admin)

        self.class_g7 = Class(name="7A", grade="七年级", school_id=self.school.id)
        self.class_g8 = Class(name="8A", grade="八年级", school_id=self.school.id)
        self.db.add_all([self.class_g7, self.class_g8])
        self.db.commit()
        self.db.refresh(self.class_g7)
        self.db.refresh(self.class_g8)

        self.subject_chinese = Subject(name="语文", code="CN", grades="七年级,八年级", school_id=self.school.id)
        self.subject_math = Subject(name="数学", code="MATH", grades="七年级,八年级", school_id=self.school.id)
        self.subject_physics = Subject(name="物理", code="PHY", grades="八年级", school_id=self.school.id)
        self.db.add_all([self.subject_chinese, self.subject_math, self.subject_physics])
        self.db.commit()
        self.db.refresh(self.subject_chinese)
        self.db.refresh(self.subject_math)
        self.db.refresh(self.subject_physics)

        self.student_g7 = Student(student_no="S7001", name="张七", gender="M", class_id=self.class_g7.id)
        self.student_g8 = Student(student_no="S8001", name="张八", gender="F", class_id=self.class_g8.id)
        self.db.add_all([self.student_g7, self.student_g8])
        self.db.commit()
        self.db.refresh(self.student_g7)
        self.db.refresh(self.student_g8)

        def override_get_db():
            db = self.SessionTesting()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[get_current_user] = lambda: self.admin
        app.dependency_overrides[require_admin_or_school_admin] = lambda: self.admin
        app.dependency_overrides[require_teacher_or_admin] = lambda: self.admin
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        self.db.close()
        self.engine.dispose()
        os.unlink(self.tmp.name)

    def _create_exam(self):
        response = self.client.post(
            "/api/exams",
            json={
                "name": "2026春季联考",
                "exam_date": "2026-04-01",
                "grade": "七年级,八年级",
                "description": "期中考试",
                "grade_subjects": [
                    {
                        "grade": "七年级",
                        "subject_ids": [self.subject_chinese.id, self.subject_math.id],
                    },
                    {
                        "grade": "八年级",
                        "subject_ids": [self.subject_chinese.id, self.subject_math.id, self.subject_physics.id],
                    },
                ],
            },
        )
        self.assertEqual(response.status_code, 201)
        return response

    def test_create_exam_returns_and_persists_grade_subjects(self):
        response = self._create_exam()
        payload = response.json()

        self.assertEqual(payload["grade_subjects"][0]["grade"], "七年级")
        self.assertEqual(payload["grade_subjects"][0]["subject_ids"], [self.subject_chinese.id, self.subject_math.id])

        with self.engine.connect() as conn:
            rows = conn.execute(
                text(
                    "SELECT grade, subject_id FROM exam_grade_subjects "
                    "WHERE exam_id = :exam_id ORDER BY grade, subject_id"
                ),
                {"exam_id": payload["id"]},
            ).fetchall()

        self.assertEqual(
            rows,
            [
                ("七年级", self.subject_chinese.id),
                ("七年级", self.subject_math.id),
                ("八年级", self.subject_chinese.id),
                ("八年级", self.subject_math.id),
                ("八年级", self.subject_physics.id),
            ],
        )

    def test_update_exam_replaces_grade_subjects(self):
        create_payload = self._create_exam().json()

        response = self.client.put(
            f"/api/exams/{create_payload['id']}",
            json={
                "grade": "七年级",
                "grade_subjects": [
                    {
                        "grade": "七年级",
                        "subject_ids": [self.subject_chinese.id],
                    }
                ],
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["grade"], "七年级")
        self.assertEqual(len(payload["grade_subjects"]), 1)
        self.assertEqual(payload["grade_subjects"][0]["subject_ids"], [self.subject_chinese.id])

        with self.engine.connect() as conn:
            rows = conn.execute(
                text(
                    "SELECT grade, subject_id FROM exam_grade_subjects "
                    "WHERE exam_id = :exam_id ORDER BY grade, subject_id"
                ),
                {"exam_id": create_payload["id"]},
            ).fetchall()

        self.assertEqual(rows, [("七年级", self.subject_chinese.id)])

    def test_score_entry_subjects_follow_student_grade_and_forbid_unconfigured_subject(self):
        exam_payload = self._create_exam().json()

        subject_response = self.client.get(
            "/api/scores/entry-subjects",
            params={"exam_id": exam_payload["id"], "student_id": self.student_g7.id},
        )
        self.assertEqual(subject_response.status_code, 200)
        self.assertEqual(
            [item["name"] for item in subject_response.json()["subjects"]],
            ["语文", "数学"],
        )

        invalid_score_response = self.client.post(
            "/api/scores",
            json={
                "student_id": self.student_g7.id,
                "exam_id": exam_payload["id"],
                "subject_id": self.subject_physics.id,
                "score": 88,
            },
        )
        self.assertEqual(invalid_score_response.status_code, 400)

        valid_score_response = self.client.post(
            "/api/scores",
            json={
                "student_id": self.student_g7.id,
                "exam_id": exam_payload["id"],
                "subject_id": self.subject_chinese.id,
                "score": 91,
            },
        )
        self.assertEqual(valid_score_response.status_code, 201)

    def test_delete_exam_with_scores_deletes_related_score_data(self):
        exam_payload = self._create_exam().json()
        exam_id = exam_payload["id"]
        self.db.add_all(
            [
                Score(student_id=self.student_g7.id, exam_id=exam_id, subject_id=self.subject_chinese.id, score=91),
                TotalRank(student_id=self.student_g7.id, exam_id=exam_id, total_score=91, rank_class=1, rank_grade=1),
            ]
        )
        self.db.commit()

        response = self.client.delete(f"/api/exams/{exam_id}")

        self.assertEqual(response.status_code, 204)
        self.assertEqual(self.db.query(Score).filter(Score.exam_id == exam_id).count(), 0)
        self.assertEqual(self.db.query(TotalRank).filter(TotalRank.exam_id == exam_id).count(), 0)
        with self.engine.connect() as conn:
            grade_subject_count = conn.execute(
                text("SELECT COUNT(*) FROM exam_grade_subjects WHERE exam_id = :exam_id"),
                {"exam_id": exam_id},
            ).scalar_one()
        self.assertEqual(grade_subject_count, 0)

    def test_export_scores_splits_grade_sheets_with_grade_specific_subject_headers(self):
        exam_payload = self._create_exam().json()
        exam_id = exam_payload["id"]

        self.db.add_all(
            [
                Score(student_id=self.student_g7.id, exam_id=exam_id, subject_id=self.subject_chinese.id, score=91),
                Score(student_id=self.student_g7.id, exam_id=exam_id, subject_id=self.subject_math.id, score=95),
                Score(student_id=self.student_g8.id, exam_id=exam_id, subject_id=self.subject_chinese.id, score=89),
                Score(student_id=self.student_g8.id, exam_id=exam_id, subject_id=self.subject_math.id, score=90),
                Score(student_id=self.student_g8.id, exam_id=exam_id, subject_id=self.subject_physics.id, score=87),
            ]
        )
        self.db.commit()

        response = self.client.get("/api/scores/export", params={"exam_id": exam_id})
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["七年级", "八年级"])

        grade7_headers = [cell.value for cell in workbook["七年级"][1]]
        grade8_headers = [cell.value for cell in workbook["八年级"][1]]
        self.assertEqual(grade7_headers, ["学号", "姓名", "班级", "语文", "数学", "总分", "班级排名", "年级排名"])
        self.assertEqual(grade8_headers, ["学号", "姓名", "班级", "语文", "数学", "物理", "总分", "班级排名", "年级排名"])

    def test_import_scores_creates_missing_student_with_unknown_gender(self):
        exam_payload = self._create_exam().json()
        exam_id = exam_payload["id"]

        wb = Workbook()
        ws = wb.active
        ws.append(["班级", "姓名", "考号", "语文", "数学"])
        ws.append([self.class_g7.name, "新学生", "S7002", 88, 92])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = self.client.post(
            "/api/scores/import",
            params={"exam_id": exam_id, "grade": "七年级"},
            files={
                "file": (
                    "scores.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success_count"], 1)
        created = self.db.query(Student).filter(Student.student_no == "S7002").first()
        self.assertIsNotNone(created)
        self.assertEqual(created.name, "新学生")
        self.assertEqual(created.gender, "U")
        self.assertEqual(created.class_id, self.class_g7.id)

        imported_scores = self.db.query(Score).filter(Score.student_id == created.id, Score.exam_id == exam_id).all()
        self.assertEqual(len(imported_scores), 2)

    def test_import_scores_fills_missing_required_subject_score_with_zero(self):
        exam_payload = self._create_exam().json()
        exam_id = exam_payload["id"]

        wb = Workbook()
        ws = wb.active
        ws.append(["班级", "姓名", "考号", "语文", "数学"])
        ws.append([self.class_g7.name, "缺分学生", "S7003", 88, None])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = self.client.post(
            "/api/scores/import",
            params={"exam_id": exam_id, "grade": "七年级"},
            files={
                "file": (
                    "scores.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success_count"], 1)
        self.assertEqual(response.json()["error_count"], 0)
        created = self.db.query(Student).filter(Student.student_no == "S7003").first()
        self.assertIsNotNone(created)

        imported_scores = {
            score.subject_id: score.score
            for score in self.db.query(Score).filter(Score.student_id == created.id, Score.exam_id == exam_id).all()
        }
        self.assertEqual(imported_scores[self.subject_chinese.id], 88)
        self.assertEqual(imported_scores[self.subject_math.id], 0)


if __name__ == "__main__":
    unittest.main()
