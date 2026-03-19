import io
import json
from datetime import date

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.dependencies import get_db, require_teacher_or_admin, get_accessible_class_ids
from app.models.student import Student
from app.models.class_ import Class
from app.models.custom_field import CustomFieldDefinition
from app.models.user import User
from app.schemas.student import StudentCreate, StudentUpdate, StudentResponse, PaginatedResponse, BatchDeleteRequest, BatchDeleteResponse

router = APIRouter(prefix="/api/students", tags=["学生管理"])


@router.get("", response_model=PaginatedResponse)
def list_students(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=9999),
    student_no: str | None = None,
    name: str | None = None,
    class_id: int | None = None,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    accessible = get_accessible_class_ids(current_user, db)
    query = db.query(Student)

    if accessible is not None:
        query = query.filter(Student.class_id.in_(accessible))
    if student_no:
        query = query.filter(Student.student_no.contains(student_no))
    if name:
        query = query.filter(Student.name.contains(name))
    if class_id:
        if accessible is not None and class_id not in accessible:
            raise HTTPException(status_code=403, detail="无权访问该班级")
        query = query.filter(Student.class_id == class_id)

    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return PaginatedResponse(
        items=[StudentResponse.model_validate(s) for s in items],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/template")
def download_template(
    _: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    headers = ["学号", "班级名称", "姓名", "性别(男/女)", "出生日期(YYYY-MM-DD)", "联系方式"]
    custom_fields = db.query(CustomFieldDefinition).order_by(CustomFieldDefinition.sort_order).all()
    for cf in custom_fields:
        headers.append(cf.field_name)
    ws.append(headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"},
    )


@router.get("/export")
def export_students(
    class_id: int | None = None,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    accessible = get_accessible_class_ids(current_user, db)
    query = db.query(Student)
    if accessible is not None:
        query = query.filter(Student.class_id.in_(accessible))
    if class_id:
        query = query.filter(Student.class_id == class_id)

    students = query.all()
    classes = {c.id: c.name for c in db.query(Class).all()}
    custom_fields = db.query(CustomFieldDefinition).order_by(CustomFieldDefinition.sort_order).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生列表"
    headers = ["学号", "班级", "姓名", "性别", "出生日期", "联系方式"]
    for cf in custom_fields:
        headers.append(cf.field_name)
    ws.append(headers)

    for s in students:
        gender_display = "男" if s.gender == "M" else "女"
        row = [
            s.student_no,
            classes.get(s.class_id, ""),
            s.name,
            gender_display,
            str(s.birth_date) if s.birth_date else "",
            s.phone or "",
        ]
        cf_data = json.loads(s.custom_fields) if s.custom_fields else {}
        for cf in custom_fields:
            row.append(cf_data.get(cf.field_name, ""))
        ws.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"},
    )


@router.post("/import")
def import_students(
    file: UploadFile = File(...),
    class_id: int | None = Query(None, description="指定班级ID，若传入则忽略Excel班级列"),
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件")

    accessible = get_accessible_class_ids(current_user, db)

    # Validate class_id if provided
    if class_id is not None:
        cls = db.query(Class).filter(Class.id == class_id).first()
        if not cls:
            raise HTTPException(status_code=400, detail="指定的班级不存在")
        if accessible is not None and class_id not in accessible:
            raise HTTPException(status_code=403, detail="无权导入到该班级")

    classes = {c.name: c.id for c in db.query(Class).all()}
    custom_fields = db.query(CustomFieldDefinition).order_by(CustomFieldDefinition.sort_order).all()
    cf_names = [cf.field_name for cf in custom_fields]

    wb = load_workbook(file.file)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success_count = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        try:
            if not row or not row[0]:
                continue
            student_no = str(row[0]).strip()
            class_name = str(row[1]).strip() if row[1] else ""
            name = str(row[2]).strip() if row[2] else ""
            gender_raw = str(row[3]).strip() if row[3] else ""
            birth_str = str(row[4]).strip() if row[4] else ""
            phone = str(row[5]).strip() if len(row) > 5 and row[5] else ""

            if not name:
                errors.append({"row": idx, "error": "姓名不能为空"})
                continue

            # Convert 男/女 to M/F
            gender_map = {"男": "M", "女": "F", "M": "M", "F": "F"}
            gender = gender_map.get(gender_raw)
            if not gender:
                errors.append({"row": idx, "error": "性别必须为 男 或 女"})
                continue

            # Determine class_id: use provided class_id or lookup from Excel
            if class_id is not None:
                target_class_id = class_id
            else:
                if class_name not in classes:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 不存在"})
                    continue
                target_class_id = classes[class_name]
                if accessible is not None and target_class_id not in accessible:
                    errors.append({"row": idx, "error": "无权导入到该班级"})
                    continue

            birth_date = None
            if birth_str:
                try:
                    birth_date = date.fromisoformat(birth_str)
                except ValueError:
                    errors.append({"row": idx, "error": "出生日期格式错误，应为 YYYY-MM-DD"})
                    continue

            existing = db.query(Student).filter(Student.student_no == student_no).first()
            if existing:
                errors.append({"row": idx, "error": f"学号 '{student_no}' 已存在"})
                continue

            cf_data = {}
            for i, cf_name in enumerate(cf_names):
                col_idx = 6 + i
                if len(row) > col_idx and row[col_idx]:
                    cf_data[cf_name] = str(row[col_idx]).strip()

            student = Student(
                student_no=student_no,
                name=name,
                gender=gender,
                birth_date=birth_date,
                class_id=target_class_id,
                phone=phone,
                custom_fields=json.dumps(cf_data, ensure_ascii=False) if cf_data else None,
            )
            db.add(student)
            db.flush()
            success_count += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    db.commit()
    return {"success_count": success_count, "error_count": len(errors), "errors": errors}


@router.get("/{student_id}", response_model=StudentResponse)
def get_student(
    student_id: int,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    accessible = get_accessible_class_ids(current_user, db)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权访问该学生")
    return student


@router.post("", response_model=StudentResponse, status_code=status.HTTP_201_CREATED)
def create_student(
    req: StudentCreate,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    accessible = get_accessible_class_ids(current_user, db)
    if accessible is not None and req.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该班级")
    cls = db.query(Class).filter(Class.id == req.class_id).first()
    if not cls:
        raise HTTPException(status_code=400, detail="班级不存在")
    existing = db.query(Student).filter(Student.student_no == req.student_no).first()
    if existing:
        raise HTTPException(status_code=400, detail="学号已存在")
    student = Student(
        student_no=req.student_no,
        name=req.name,
        gender=req.gender,
        birth_date=req.birth_date,
        class_id=req.class_id,
        phone=req.phone,
        custom_fields=req.custom_fields,
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return student


@router.put("/{student_id}", response_model=StudentResponse)
def update_student(
    student_id: int,
    req: StudentUpdate,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    accessible = get_accessible_class_ids(current_user, db)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该学生")

    if req.student_no is not None:
        dup = db.query(Student).filter(Student.student_no == req.student_no, Student.id != student_id).first()
        if dup:
            raise HTTPException(status_code=400, detail="学号已存在")
        student.student_no = req.student_no
    if req.name is not None:
        student.name = req.name
    if req.gender is not None:
        student.gender = req.gender
    if req.birth_date is not None:
        student.birth_date = req.birth_date
    if req.class_id is not None:
        if accessible is not None and req.class_id not in accessible:
            raise HTTPException(status_code=403, detail="无权操作该班级")
        student.class_id = req.class_id
    if req.phone is not None:
        student.phone = req.phone
    if req.custom_fields is not None:
        student.custom_fields = req.custom_fields

    db.commit()
    db.refresh(student)
    return student


@router.delete("/batch", response_model=BatchDeleteResponse)
def batch_delete_students(
    req: BatchDeleteRequest,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    if not req.ids:
        raise HTTPException(status_code=400, detail="请选择要删除的学生")
    accessible = get_accessible_class_ids(current_user, db)
    students = db.query(Student).filter(Student.id.in_(req.ids)).all()
    deleted_count = 0
    for student in students:
        if accessible is not None and student.class_id not in accessible:
            continue
        db.delete(student)
        deleted_count += 1
    db.commit()
    return BatchDeleteResponse(deleted_count=deleted_count)


@router.delete("/{student_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_student(
    student_id: int,
    _: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")
    db.delete(student)
    db.commit()
