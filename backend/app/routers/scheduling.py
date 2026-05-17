from __future__ import annotations

import io
import json
import re
from datetime import datetime
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import MetaData, Table
from sqlalchemy.orm import Session, sessionmaker

from app.database import SessionLocal
from app.dependencies import get_db, require_school_admin
from app.models.class_ import Class
from app.models.class_timetable import ClassTimetable
from app.models.lesson_plan import LessonPlan
from app.models.lesson_plan_override import LessonPlanOverride
from app.models.schedule_draft import ScheduleDraft
from app.models.schedule_draft_item import ScheduleDraftItem
from app.models.schedule_import import ScheduleImport
from app.models.schedule_import_item import ScheduleImportItem
from app.models.schedule_period import SchedulePeriod
from app.models.schedule_period_plan import SchedulePeriodPlan
from app.models.schedule_task import ScheduleTask
from app.models.subject import Subject
from app.models.teacher_class_subject import TeacherClassSubject
from app.models.teacher_time_constraint import TeacherTimeConstraint
from app.models.timetable_lock import TimetableLock
from app.models.user import User
from app.schemas.scheduling import (
    DraftDiagnostic,
    DraftItem,
    DraftPublishCheck,
    DraftSummaryPayload,
    LessonPlanBatchResponse,
    LessonPlanBatchSaveRequest,
    LessonPlanConfig,
    LessonPlanOverrideBatchResponse,
    LessonPlanOverrideBatchSaveRequest,
    LessonPlanOverrideItem,
    PeriodPlanResponse,
    PeriodPlanSaveRequest,
    PublishDraftResponse,
    ScheduleDraftItemsResponse,
    ScheduleDraftResponse,
    ScheduleImportResponse,
    ScheduleImportItemsResponse,
    ScheduleImportItemResponse,
    ScheduleImportItemUpdate,
    ScheduleImportDraftCreateResponse,
    ScheduleImportSummary,
    ScheduleTaskCreateResponse,
    ScheduleTaskResponse,
    TeacherConstraintBatchResponse,
    TeacherConstraintBatchSaveRequest,
    TeacherConstraintItem,
    TeachingArrangementBatchResponse,
    TeachingArrangementBatchSaveRequest,
    TeachingArrangementItem,
    TimetableLockBatchResponse,
    TimetableLockBatchSaveRequest,
    TimetableLockItem,
)
from app.services.scheduling.compiler import compile_problem
from app.services.scheduling.config_loader import decode_json_content, decode_period_plan_ids, load_scheduling_raw_config
from app.services.scheduling.cp_sat_solver import solve_schedule
from app.services.scheduling.debug_export import build_scheduling_debug_package
from app.services.scheduling.draft_service import create_draft_from_solution, serialize_draft, serialize_draft_items
from app.services.scheduling.import_service import build_schedule_import_template, create_draft_from_import, create_import_from_upload, serialize_import_item, update_import_item
from app.services.scheduling.publish_service import publish_draft
from app.services.scheduling.validators import validate_compiled_problem, validate_raw_config

router = APIRouter(prefix="/api/schedule", tags=["schedule"])


def _require_school_id(current_user: User) -> int:
    if current_user.school_id is None:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="当前账户缺少学校信息")
    return current_user.school_id


def _json_dumps(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False)


def _to_task_response(task: ScheduleTask) -> ScheduleTaskResponse:
    result: Any = None
    if task.result:
        try:
            result = json.loads(task.result)
        except json.JSONDecodeError:
            result = task.result
    return ScheduleTaskResponse(
        id=task.id,
        status=task.status,
        progress=task.progress,
        message=task.message,
        result=result,
        error=task.error,
        started_at=task.started_at,
        finished_at=task.finished_at,
    )


def _decode_import_summary(row: ScheduleImport) -> ScheduleImportSummary:
    if not row.summary:
        return ScheduleImportSummary()
    try:
        payload = json.loads(row.summary)
    except json.JSONDecodeError:
        return ScheduleImportSummary()
    return ScheduleImportSummary(**payload)


def _to_import_response(row: ScheduleImport) -> ScheduleImportResponse:
    return ScheduleImportResponse(
        id=row.id,
        grade=row.grade,
        scope=row.scope,
        class_id=row.class_id,
        source_type=row.source_type,
        status=row.status,
        message=row.message,
        error=row.error,
        summary=_decode_import_summary(row),
        created_at=row.created_at,
        updated_at=row.updated_at,
    )


def _encode_lesson_plan(item: LessonPlanConfig | LessonPlanOverrideItem | TeacherConstraintItem) -> str:
    payload = item.model_dump()
    payload.pop("subject_id", None)
    payload.pop("class_id", None)
    payload.pop("teacher_id", None)
    return _json_dumps(payload)


def _available_auto_period_ids(db: Session, school_id: int) -> list[int]:
    rows = (
        db.query(SchedulePeriod)
        .filter(
            SchedulePeriod.school_id == school_id,
            SchedulePeriod.is_active == True,
            SchedulePeriod.include_in_auto_schedule == True,
        )
        .order_by(SchedulePeriod.sort_order, SchedulePeriod.id)
        .all()
    )
    return [int(row.id) for row in rows]


def _normalize_period_ids(period_ids: list[int], available_ids: list[int]) -> list[int]:
    available = set(available_ids)
    result: list[int] = []
    seen: set[int] = set()
    for raw_id in period_ids:
        period_id = int(raw_id)
        if period_id in seen:
            continue
        if period_id not in available:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="节次计划包含不可用节次")
        result.append(period_id)
        seen.add(period_id)
    return result


def _build_lesson_plan_row_values(grade: str, school_id: int, item: LessonPlanConfig, lesson_plan_table: Table, *, include_timestamps: bool) -> dict[str, Any]:
    now = datetime.now()
    values: dict[str, Any] = {
        "school_id": school_id,
        "grade": grade,
        "subject_id": item.subject_id,
        "content": _encode_lesson_plan(item),
    }
    columns = lesson_plan_table.c.keys()
    if "weekly_hours" in columns:
        values["weekly_hours"] = item.weekly_hours
    if "priority" in columns:
        values["priority"] = 1
    if "avoid_consecutive" in columns:
        values["avoid_consecutive"] = False
    if "forbidden_periods_json" in columns:
        values["forbidden_periods_json"] = _json_dumps(item.forbidden_periods)
    if include_timestamps and "created_at" in columns:
        values["created_at"] = now
    if "updated_at" in columns:
        values["updated_at"] = now
    return values


def _build_timetable_names(db: Session, rows: list[ScheduleDraftItem]) -> tuple[dict[int, str], dict[int, str], dict[int, str], dict[int, str]]:
    class_ids = sorted({row.class_id for row in rows})
    subject_ids = sorted({row.subject_id for row in rows})
    teacher_ids = sorted({row.teacher_id for row in rows})
    period_ids = sorted({row.period_id for row in rows})
    class_name_map = {item.id: item.name for item in db.query(Class).filter(Class.id.in_(class_ids)).all()} if class_ids else {}
    subject_name_map = {item.id: item.name for item in db.query(Subject).filter(Subject.id.in_(subject_ids)).all()} if subject_ids else {}
    teacher_name_map = {item.id: item.username for item in db.query(User).filter(User.id.in_(teacher_ids)).all()} if teacher_ids else {}
    period_name_map = {item.id: item.name for item in db.query(SchedulePeriod).filter(SchedulePeriod.id.in_(period_ids)).all()} if period_ids else {}
    return class_name_map, subject_name_map, teacher_name_map, period_name_map


def _safe_sheet_title(value: str, fallback: str, used_titles: set[str]) -> str:
    title = "".join(ch for ch in (value or fallback) if ch not in r'[]:*?/\\').strip() or fallback
    title = title[:31]
    if title not in used_titles:
        used_titles.add(title)
        return title
    base = title[:28]
    index = 2
    while True:
        candidate = f"{base}-{index}"[:31]
        if candidate not in used_titles:
            used_titles.add(candidate)
            return candidate
        index += 1


_CHINESE_DIGITS = {
    "一": 1,
    "二": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
}


def _parse_standard_chinese_number(value: str) -> int | None:
    if not value:
        return None
    if value == "十":
        return 10
    if "十" in value:
        left, right = value.split("十", 1)
        tens = _CHINESE_DIGITS.get(left, 1) if left else 1
        ones = _CHINESE_DIGITS.get(right, 0) if right else 0
        return tens * 10 + ones
    return _CHINESE_DIGITS.get(value)


def _extract_class_number(class_name: str) -> int | None:
    name = (class_name or "").strip()
    match = re.search(r"(\d+)\s*班$", name)
    if match:
        return int(match.group(1))
    match = re.search(r"([一二三四五六七八九十]+)\s*班$", name)
    if not match:
        return None
    token = match.group(1)
    if len(token) > 1 and token[0] in "七八九":
        suffix_number = _parse_standard_chinese_number(token[1:])
        if suffix_number is not None:
            return suffix_number
    return _parse_standard_chinese_number(token)


def _class_sheet_sort_key(class_id: int, class_name_map: dict[int, str]) -> tuple[Any, ...]:
    class_name = class_name_map.get(class_id, str(class_id))
    class_number = _extract_class_number(class_name)
    if class_number is not None:
        return (0, class_number, class_name, class_id)
    return (1, class_name, class_id)


def _is_afternoon_period(period: SchedulePeriod) -> bool:
    start_time = str(period.start_time or "")
    if not start_time:
        return False
    hour_text = start_time.split(":", 1)[0]
    return hour_text.isdigit() and int(hour_text) >= 12


def _export_periods_for_draft(db: Session, school_id: int, grade: str, rows: list[ScheduleDraftItem]) -> list[SchedulePeriod]:
    period_plan = (
        db.query(SchedulePeriodPlan)
        .filter(SchedulePeriodPlan.school_id == school_id, SchedulePeriodPlan.grade == grade)
        .first()
    )
    if period_plan:
        selected_period_ids = decode_period_plan_ids(period_plan.config)
        if not selected_period_ids:
            return []
        return (
            db.query(SchedulePeriod)
            .filter(
                SchedulePeriod.school_id == school_id,
                SchedulePeriod.is_active == True,  # noqa: E712
                SchedulePeriod.include_in_auto_schedule == True,  # noqa: E712
                SchedulePeriod.id.in_(selected_period_ids),
            )
            .order_by(SchedulePeriod.sort_order, SchedulePeriod.id)
            .all()
        )

    configured_periods = (
        db.query(SchedulePeriod)
        .filter(
            SchedulePeriod.school_id == school_id,
            SchedulePeriod.is_active == True,  # noqa: E712
            SchedulePeriod.include_in_auto_schedule == True,  # noqa: E712
        )
        .order_by(SchedulePeriod.sort_order, SchedulePeriod.id)
        .all()
    )
    if configured_periods:
        return configured_periods

    period_ids = sorted({row.period_id for row in rows})
    return db.query(SchedulePeriod).filter(SchedulePeriod.id.in_(period_ids)).order_by(SchedulePeriod.sort_order, SchedulePeriod.id).all() if period_ids else []


def _style_schedule_sheet(ws, max_row: int, max_col: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    ws.column_dimensions["A"].width = 12
    for column_index in range(2, max_col + 1):
        ws.column_dimensions[chr(64 + column_index)].width = 14
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 26

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F7F7F7")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"].font = Font(name="SimSun", size=20, bold=True)
    ws["A1"].alignment = center
    ws["A2"].font = Font(name="SimSun", size=14, bold=True)
    ws["A2"].alignment = center

    for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = center
            cell.font = Font(name="SimSun", size=11, bold=True)
            if cell.row == 3:
                cell.fill = header_fill
    ws.freeze_panes = "B4"


def _mark_task_failed(db: Session, task: ScheduleTask, message: str, *, error: str | None = None, result: dict[str, Any] | None = None) -> None:
    task.status = "failed"
    task.progress = 100
    task.message = message
    task.error = error or message
    task.result = _json_dumps(result or {}) if result is not None else None
    task.finished_at = datetime.now()
    db.commit()


@router.get("/lesson-plan/{grade}", response_model=LessonPlanBatchResponse)
def get_lesson_plan(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    rows = db.query(LessonPlan).filter(LessonPlan.school_id == school_id, LessonPlan.grade == grade).order_by(LessonPlan.subject_id).all()
    items = []
    for row in rows:
        payload = decode_json_content(row.content)
        items.append(LessonPlanConfig(subject_id=row.subject_id, **payload))
    return LessonPlanBatchResponse(grade=grade, items=items)


@router.get("/period-plan/{grade}", response_model=PeriodPlanResponse)
def get_period_plan(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    row = (
        db.query(SchedulePeriodPlan)
        .filter(SchedulePeriodPlan.school_id == school_id, SchedulePeriodPlan.grade == grade)
        .first()
    )
    period_ids = decode_period_plan_ids(row.config) if row else _available_auto_period_ids(db, school_id)
    return PeriodPlanResponse(grade=grade, period_ids=period_ids)


@router.post("/period-plan", response_model=PeriodPlanResponse)
def save_period_plan(req: PeriodPlanSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    available_ids = _available_auto_period_ids(db, school_id)
    period_ids = _normalize_period_ids(req.period_ids, available_ids)
    row = (
        db.query(SchedulePeriodPlan)
        .filter(SchedulePeriodPlan.school_id == school_id, SchedulePeriodPlan.grade == req.grade)
        .first()
    )
    if row is None:
        row = SchedulePeriodPlan(school_id=school_id, grade=req.grade)
        db.add(row)
    row.config = _json_dumps({"period_ids": period_ids})
    db.commit()
    return PeriodPlanResponse(grade=req.grade, period_ids=period_ids)


@router.post("/lesson-plan", response_model=LessonPlanBatchResponse)
def save_lesson_plan(req: LessonPlanBatchSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    existing = {
        row.subject_id: row
        for row in db.query(LessonPlan).filter(LessonPlan.school_id == school_id, LessonPlan.grade == req.grade).all()
    }
    lesson_plan_table = Table("lesson_plans", MetaData(), autoload_with=db.get_bind())
    for item in req.items:
        row = existing.get(item.subject_id)
        values = _build_lesson_plan_row_values(req.grade, school_id, item, lesson_plan_table, include_timestamps=row is None)
        if row is None:
            db.execute(lesson_plan_table.insert().values(**values))
        else:
            db.execute(lesson_plan_table.update().where(lesson_plan_table.c.id == row.id).values(**values))
    db.commit()
    return req


@router.get("/teaching-arrangement/{grade}", response_model=TeachingArrangementBatchResponse)
def get_teaching_arrangement(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    class_ids = [row.id for row in db.query(Class).filter(Class.school_id == school_id, Class.grade == grade).all()]
    rows = db.query(TeacherClassSubject).filter(TeacherClassSubject.school_id == school_id, TeacherClassSubject.class_id.in_(class_ids)).all() if class_ids else []
    items = [TeachingArrangementItem(class_id=row.class_id, subject_id=row.subject_id, teacher_id=row.teacher_id) for row in rows]
    return TeachingArrangementBatchResponse(grade=grade, items=items)


@router.post("/teaching-arrangement", response_model=TeachingArrangementBatchResponse)
def save_teaching_arrangement(req: TeachingArrangementBatchSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    class_ids = [row.id for row in db.query(Class).filter(Class.school_id == school_id, Class.grade == req.grade).all()]
    if class_ids:
        db.query(TeacherClassSubject).filter(TeacherClassSubject.school_id == school_id, TeacherClassSubject.class_id.in_(class_ids)).delete(synchronize_session=False)
    for item in req.items:
        db.add(TeacherClassSubject(school_id=school_id, class_id=item.class_id, subject_id=item.subject_id, teacher_id=item.teacher_id))
    db.commit()
    return TeachingArrangementBatchResponse(grade=req.grade, items=req.items)


@router.get("/lesson-plan-overrides/{grade}", response_model=LessonPlanOverrideBatchResponse)
def get_lesson_plan_overrides(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    rows = db.query(LessonPlanOverride).filter(LessonPlanOverride.school_id == school_id, LessonPlanOverride.grade == grade).all()
    items = []
    for row in rows:
        payload = decode_json_content(row.config)
        items.append(LessonPlanOverrideItem(class_id=row.class_id, subject_id=row.subject_id, **payload))
    return LessonPlanOverrideBatchResponse(grade=grade, items=items)


@router.post("/lesson-plan-overrides", response_model=LessonPlanOverrideBatchResponse)
def save_lesson_plan_overrides(req: LessonPlanOverrideBatchSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    db.query(LessonPlanOverride).filter(LessonPlanOverride.school_id == school_id, LessonPlanOverride.grade == req.grade).delete(synchronize_session=False)
    for item in req.items:
        db.add(LessonPlanOverride(school_id=school_id, grade=req.grade, class_id=item.class_id, subject_id=item.subject_id, config=_encode_lesson_plan(item)))
    db.commit()
    return LessonPlanOverrideBatchResponse(grade=req.grade, items=req.items)


@router.get("/teacher-constraints/{grade}", response_model=TeacherConstraintBatchResponse)
def get_teacher_constraints(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    rows = db.query(TeacherTimeConstraint).filter(TeacherTimeConstraint.school_id == school_id, TeacherTimeConstraint.grade == grade).all()
    items = []
    for row in rows:
        payload = decode_json_content(row.config)
        items.append(TeacherConstraintItem(teacher_id=row.teacher_id, **payload))
    return TeacherConstraintBatchResponse(grade=grade, items=items)


@router.post("/teacher-constraints", response_model=TeacherConstraintBatchResponse)
def save_teacher_constraints(req: TeacherConstraintBatchSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    db.query(TeacherTimeConstraint).filter(TeacherTimeConstraint.school_id == school_id, TeacherTimeConstraint.grade == req.grade).delete(synchronize_session=False)
    for item in req.items:
        db.add(TeacherTimeConstraint(school_id=school_id, grade=req.grade, teacher_id=item.teacher_id, config=_encode_lesson_plan(item)))
    db.commit()
    return TeacherConstraintBatchResponse(grade=req.grade, items=req.items)


@router.get("/locks/{grade}", response_model=TimetableLockBatchResponse)
def get_locks(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    rows = db.query(TimetableLock).filter(TimetableLock.school_id == school_id, TimetableLock.grade == grade).all()
    items = [
        TimetableLockItem(
            class_id=row.class_id,
            subject_id=row.subject_id,
            teacher_id=row.teacher_id,
            weekday=row.weekday,
            period_id=row.period_id,
            source=row.source or "manual",
            note=row.note or "",
        )
        for row in rows
    ]
    return TimetableLockBatchResponse(grade=grade, items=items)


@router.post("/locks", response_model=TimetableLockBatchResponse)
def save_locks(req: TimetableLockBatchSaveRequest, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    db.query(TimetableLock).filter(TimetableLock.school_id == school_id, TimetableLock.grade == req.grade).delete(synchronize_session=False)
    for item in req.items:
        db.add(
            TimetableLock(
                school_id=school_id,
                grade=req.grade,
                class_id=item.class_id,
                subject_id=item.subject_id,
                teacher_id=item.teacher_id,
                weekday=item.weekday,
                period_id=item.period_id,
                source=item.source,
                note=item.note,
                created_by=current_user.id,
            )
        )
    db.commit()
    return TimetableLockBatchResponse(grade=req.grade, items=req.items)


@router.get("/debug-config/{grade}/export")
def export_schedule_debug_config(grade: str, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    raw_config = load_scheduling_raw_config(db, school_id, grade)
    exported_at = datetime.now().isoformat()
    payload = build_scheduling_debug_package(raw_config, grade=grade, exported_at=exported_at)
    filename = f"schedule-debug-{grade}-{datetime.now().strftime('%Y%m%d%H%M%S')}.json"
    return Response(
        content=_json_dumps(payload),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/imports/template")
def download_schedule_import_template(current_user: User = Depends(require_school_admin)):
    _require_school_id(current_user)
    template = build_schedule_import_template()
    headers = {"Content-Disposition": 'attachment; filename="schedule-import-template.xlsx"'}
    return StreamingResponse(
        template,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/imports/{import_id}", response_model=ScheduleImportResponse)
def get_schedule_import(import_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    row = db.query(ScheduleImport).filter(ScheduleImport.id == import_id, ScheduleImport.school_id == school_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import not found")
    return _to_import_response(row)


@router.post("/imports", response_model=ScheduleImportResponse, status_code=201)
def create_schedule_import(
    grade: str = Form(...),
    scope: str = Form(...),
    class_id: int | None = Form(None),
    file: UploadFile = File(...),
    current_user: User = Depends(require_school_admin),
    db: Session = Depends(get_db),
):
    school_id = _require_school_id(current_user)
    row = create_import_from_upload(
        db,
        school_id=school_id,
        grade=grade,
        scope=scope,
        class_id=class_id,
        file=file,
        created_by=current_user.id,
    )
    return _to_import_response(row)


@router.get("/imports/{import_id}/items", response_model=ScheduleImportItemsResponse)
def get_schedule_import_items(import_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    row = db.query(ScheduleImport).filter(ScheduleImport.id == import_id, ScheduleImport.school_id == school_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import not found")
    items = db.query(ScheduleImportItem).filter(ScheduleImportItem.import_id == row.id).order_by(ScheduleImportItem.class_id, ScheduleImportItem.weekday, ScheduleImportItem.period_id).all()
    return ScheduleImportItemsResponse(items=[ScheduleImportItemResponse(**serialize_import_item(db, item)) for item in items])


@router.patch("/imports/{import_id}/items/{item_id}", response_model=ScheduleImportItemResponse)
def patch_schedule_import_item(
    import_id: int,
    item_id: int,
    req: ScheduleImportItemUpdate,
    current_user: User = Depends(require_school_admin),
    db: Session = Depends(get_db),
):
    school_id = _require_school_id(current_user)
    row = db.query(ScheduleImport).filter(ScheduleImport.id == import_id, ScheduleImport.school_id == school_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import not found")
    item = db.query(ScheduleImportItem).filter(ScheduleImportItem.id == item_id, ScheduleImportItem.import_id == row.id).first()
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import item not found")
    item = update_import_item(db, item, subject_id=req.subject_id, teacher_id=req.teacher_id, is_empty=req.is_empty)
    return ScheduleImportItemResponse(**serialize_import_item(db, item))


@router.post("/imports/{import_id}/draft", response_model=ScheduleImportDraftCreateResponse, status_code=201)
def create_schedule_import_draft(import_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    row = db.query(ScheduleImport).filter(ScheduleImport.id == import_id, ScheduleImport.school_id == school_id).first()
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="import not found")
    draft = create_draft_from_import(db, row, created_by=current_user.id)
    return ScheduleImportDraftCreateResponse(import_id=row.id, draft_id=draft.id, status=row.status)


def _run_draft_task(task_id: int, session_factory: sessionmaker = SessionLocal) -> None:
    db = session_factory()
    try:
        task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id).first()
        if not task:
            return
        task.status = "running"
        task.progress = 20
        task.message = "正在加载排课配置"
        task.started_at = datetime.now()
        task.error = None
        db.commit()

        raw_config = load_scheduling_raw_config(db, task.school_id, task.grade)
        diagnostics = validate_raw_config(raw_config)
        blocking = [item for item in diagnostics if item.get("blocking", True)]
        if blocking:
            _mark_task_failed(db, task, "排课配置校验失败", result={"diagnostics": diagnostics})
            return

        task.progress = 40
        task.message = "正在编译排课模型"
        db.commit()
        compiled = compile_problem(raw_config)
        diagnostics.extend(validate_compiled_problem(compiled))
        blocking = [item for item in diagnostics if item.get("blocking", True)]
        if blocking:
            _mark_task_failed(db, task, "排课模型校验失败", result={"diagnostics": diagnostics})
            return

        task.progress = 70
        task.message = "CP-SAT 正在求解"
        db.commit()
        result = solve_schedule(compiled)
        if not result.success:
            _mark_task_failed(db, task, "排课求解失败", result={"diagnostics": result.diagnostics})
            return

        draft = create_draft_from_solution(db, school_id=task.school_id, grade=task.grade, created_by=None, problem=compiled, result=result)
        task.status = "success"
        task.progress = 100
        task.message = "草案生成成功"
        task.result = _json_dumps({"draft_id": draft.id, "score": draft.score, "summary": json.loads(draft.summary or "{}")})
        task.finished_at = datetime.now()
        db.commit()
    except Exception as exc:
        task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id).first()
        if task:
            _mark_task_failed(db, task, "排课过程中出现异常", error=str(exc))
    finally:
        db.close()


@router.post("/drafts/{grade}/solve", response_model=ScheduleTaskCreateResponse, status_code=202)
def solve_schedule_draft(grade: str, background_tasks: BackgroundTasks, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    task = ScheduleTask(school_id=school_id, grade=grade, status="pending", progress=0, message="等待排课开始")
    db.add(task)
    db.commit()
    db.refresh(task)
    background_tasks.add_task(_run_draft_task, task.id, SessionLocal)
    return ScheduleTaskCreateResponse(task_id=task.id, status=task.status)


@router.get("/tasks/{task_id}", response_model=ScheduleTaskResponse)
def get_schedule_task(task_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id, ScheduleTask.school_id == school_id).first()
    if not task:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="task not found")
    return _to_task_response(task)


@router.get("/drafts/{draft_id}", response_model=ScheduleDraftResponse)
def get_schedule_draft(draft_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    draft = db.query(ScheduleDraft).filter(ScheduleDraft.id == draft_id, ScheduleDraft.school_id == school_id).first()
    if not draft:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="draft not found")
    payload = serialize_draft(draft)
    return ScheduleDraftResponse(
        id=payload["id"],
        grade=payload["grade"],
        status=payload["status"],
        score=payload["score"],
        summary=DraftSummaryPayload(**payload["summary"]),
        diagnostics=[DraftDiagnostic(**item) for item in payload["diagnostics"]],
        publish_checks=[DraftPublishCheck(**item) for item in payload["publish_checks"]],
        published_at=datetime.fromisoformat(payload["published_at"]) if payload["published_at"] else None,
        created_at=datetime.fromisoformat(payload["created_at"]) if payload["created_at"] else None,
        updated_at=datetime.fromisoformat(payload["updated_at"]) if payload["updated_at"] else None,
    )


@router.get("/drafts/{draft_id}/items", response_model=ScheduleDraftItemsResponse)
def get_schedule_draft_items(draft_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    draft = db.query(ScheduleDraft).filter(ScheduleDraft.id == draft_id, ScheduleDraft.school_id == school_id).first()
    if not draft:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="draft not found")
    rows = db.query(ScheduleDraftItem).filter(ScheduleDraftItem.draft_id == draft.id).order_by(ScheduleDraftItem.class_id, ScheduleDraftItem.weekday, ScheduleDraftItem.period_id).all()
    class_name_map, subject_name_map, teacher_name_map, period_name_map = _build_timetable_names(db, rows)
    base_items = serialize_draft_items(rows)
    items = [
        DraftItem(
            weekday=item["weekday"],
            period_id=item["period_id"],
            period_name=period_name_map.get(item["period_id"]),
            class_id=item["class_id"],
            class_name=class_name_map.get(item["class_id"]),
            subject_id=item["subject_id"],
            subject_name=subject_name_map.get(item["subject_id"]),
            teacher_id=item["teacher_id"],
            teacher_name=teacher_name_map.get(item["teacher_id"]),
            is_locked=item["is_locked"],
            penalty_tags=item["penalty_tags"],
        )
        for item in base_items
    ]
    return ScheduleDraftItemsResponse(items=items)


@router.get("/drafts/{draft_id}/export")
def export_schedule_draft(draft_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    draft = db.query(ScheduleDraft).filter(ScheduleDraft.id == draft_id, ScheduleDraft.school_id == school_id).first()
    if not draft:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="draft not found")

    rows = (
        db.query(ScheduleDraftItem)
        .filter(ScheduleDraftItem.draft_id == draft.id)
        .order_by(ScheduleDraftItem.class_id, ScheduleDraftItem.weekday, ScheduleDraftItem.period_id)
        .all()
    )
    class_name_map, subject_name_map, _teacher_name_map, _period_name_map = _build_timetable_names(db, rows)
    periods = _export_periods_for_draft(db, school_id, draft.grade, rows)
    class_ids = sorted({row.class_id for row in rows}, key=lambda class_id: _class_sheet_sort_key(class_id, class_name_map))

    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()
    weekdays = ["星期一", "星期二", "星期三", "星期四", "星期五"]
    for class_id in class_ids:
        class_title = class_name_map.get(class_id, "") or f"班级{class_id}"
        ws = wb.create_sheet(title=_safe_sheet_title(class_title, f"班级{class_id}", used_titles))
        ws.append(["课 程 表", "", "", "", "", ""])
        ws.append([class_title, "", "", "", "", ""])
        ws.append(["节次", *weekdays])
        class_rows = [row for row in rows if row.class_id == class_id]
        item_map = {(row.weekday, row.period_id): row for row in class_rows}
        lunch_inserted = False
        for period in periods:
            if not lunch_inserted and _is_afternoon_period(period):
                ws.append(["午休", "", "", "", "", ""])
                lunch_row = ws.max_row
                ws.merge_cells(start_row=lunch_row, start_column=1, end_row=lunch_row, end_column=6)
                lunch_inserted = True
            line = [period.name]
            for weekday in range(1, 6):
                item = item_map.get((weekday, period.id))
                line.append(subject_name_map.get(item.subject_id, "") if item else "")
            ws.append(line)
        _style_schedule_sheet(ws, ws.max_row)

    if not wb.sheetnames:
        ws = wb.create_sheet(title="课表")
        ws.append(["课 程 表", "", "", "", "", ""])
        ws.append(["课表", "", "", "", "", ""])
        ws.append(["节次", *weekdays])
        _style_schedule_sheet(ws, ws.max_row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"schedule-draft-{draft.grade}-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/drafts/{draft_id}/publish", response_model=PublishDraftResponse)
def publish_schedule_draft(draft_id: int, current_user: User = Depends(require_school_admin), db: Session = Depends(get_db)):
    school_id = _require_school_id(current_user)
    draft = db.query(ScheduleDraft).filter(ScheduleDraft.id == draft_id, ScheduleDraft.school_id == school_id).first()
    if not draft:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="draft not found")
    payload = serialize_draft(draft)
    blocking = [item for item in payload["publish_checks"] if item.get("blocking", True)]
    if blocking:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="draft has blocking checks")
    row_count = publish_draft(db, draft)
    return PublishDraftResponse(draft_id=draft.id, status=draft.status, rows=row_count)
