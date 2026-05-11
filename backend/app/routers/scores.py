import io
import json
import re

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.dependencies import get_db, get_current_user, require_teacher_or_admin, get_accessible_class_ids, get_user_school_id
from app.models.score import Score
from app.models.student import Student
from app.models.class_ import Class
from app.models.exam import Exam
from app.models.subject import Subject
from app.models.total_rank import TotalRank
from app.models.user import User
from app.schemas.score import ScoreCreate, ScoreUpdate, ScoreItemResponse, ScorePaginatedResponse, BatchDeleteScoresRequest
from app.services.exam_grade_subjects import load_exam_subjects_for_grade, parse_grade_tokens
from app.utils.ranking import recalculate_ranks
from app.routers.analysis import _load_subject_full_score_config, _subject_full_score_for_three_rates

router = APIRouter(prefix="/api/scores", tags=["成绩管理"])

# Fixed subject display order and required subjects for import
SUBJECT_DISPLAY_ORDER = ["语文", "数学", "英语", "物理", "生物", "历史", "地理", "道法", "政治", "化学"]
REQUIRED_SUBJECTS = {"语文", "数学", "英语", "历史", "道法", "政治", "思想品德"}

HEADER_ALIAS_GROUPS = {
    "学号": {"学号", "学生学号", "学籍号", "考号", "唯一号"},
    "姓名": {"姓名", "学生姓名"},
    "班级": {"班级", "班级名称"},
    "总分": {"总分", "总成绩", "总分数"},
    "班级排名": {"班级排名", "总分班名", "班名"},
    "年级排名": {"年级排名", "总分校名", "校名"},
}

SUBJECT_ALIAS_GROUPS = {
    "语文": {"语文"},
    "数学": {"数学"},
    "英语": {"英语"},
    "物理": {"物理"},
    "生物": {"生物"},
    "历史": {"历史"},
    "地理": {"地理"},
    "道法": {"道法", "政治", "道法(政治)", "思想品德", "思品"},
    "化学": {"化学"},
}

CHINESE_DIGIT_MAP = str.maketrans({
    "零": "0",
    "一": "1",
    "二": "2",
    "三": "3",
    "四": "4",
    "五": "5",
    "六": "6",
    "七": "7",
    "八": "8",
    "九": "9",
})


def _normalize_header_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("　", "")
    text = re.sub(r"\s+", "", text)
    text = text.replace("(必填)", "").replace("（必填）", "")
    return text


def _normalize_class_text(value) -> str:
    text = _normalize_header_text(value)
    return text.translate(CHINESE_DIGIT_MAP)


def _normalize_cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_class_value(value) -> bool:
    normalized = _normalize_class_text(value)
    if not normalized:
        return False
    normalized = normalized.replace("班级", "班")

    if re.fullmatch(r"\d{1,2}", normalized):
        return True
    if re.fullmatch(r"\d{1,2}班", normalized):
        return True
    if re.fullmatch(r"[高七八九]?\d{1,2}(?:\(\d+\))?班", normalized):
        return True
    return False


def _expand_class_candidates(normalized_class_name: str) -> list[str]:
    if not normalized_class_name:
        return []

    base = normalized_class_name.replace("班级", "班")
    candidates = {base}

    if re.fullmatch(r"\d+\.0+", base):
        base = str(int(float(base)))
        candidates.add(base)

    if base.endswith("班"):
        prefix = base[:-1]
        if re.fullmatch(r"\d+", prefix):
            trimmed = prefix.lstrip("0") or "0"
            candidates.update({f"{trimmed}班", trimmed})
    elif re.fullmatch(r"\d+", base):
        trimmed = base.lstrip("0") or "0"
        candidates.update({trimmed, f"{trimmed}班"})
        if len(trimmed) == 1:
            candidates.add(f"{trimmed.zfill(2)}班")

    return list(candidates)


def _extract_class_number(normalized_class_name: str) -> str | None:
    text = normalized_class_name.replace("班级", "班")
    match = re.search(r"\(?(\d{1,2})\)?班$", text)
    if not match:
        match = re.search(r"\(?(\d{1,2})\)?$", text)
    if not match:
        return None
    return str(int(match.group(1)))


def _resolve_class_ids_by_name(excel_class_name: str, class_name_map: dict[int, str]) -> list[int]:
    normalized = _normalize_class_text(excel_class_name)
    candidates = _expand_class_candidates(normalized)
    if not candidates:
        return []

    normalized_name_map = {
        cid: _normalize_class_text(name).replace("班级", "班")
        for cid, name in class_name_map.items()
    }

    exact = [cid for cid, norm_name in normalized_name_map.items() if norm_name in candidates]
    if exact:
        return exact

    suffix_match = [
        cid
        for cid, norm_name in normalized_name_map.items()
        if any(norm_name.endswith(candidate) or candidate.endswith(norm_name) for candidate in candidates)
    ]
    # Keep order stable while deduplicating.
    if suffix_match:
        return list(dict.fromkeys(suffix_match))

    candidate_numbers = {
        number
        for candidate in candidates
        for number in [_extract_class_number(candidate)]
        if number is not None
    }
    number_match = [
        cid
        for cid, norm_name in normalized_name_map.items()
        if _extract_class_number(norm_name) in candidate_numbers
    ]
    return list(dict.fromkeys(number_match))


def _should_use_explicit_ranks(
    students_with_explicit_ranks: dict[int, dict[str, int | None]],
    success_count: int,
) -> bool:
    if not students_with_explicit_ranks or success_count <= 0:
        return False
    if len(students_with_explicit_ranks) != success_count:
        return False
    return all(
        ranks.get("rank_class") is not None and ranks.get("rank_grade") is not None
        for ranks in students_with_explicit_ranks.values()
    )


def _infer_class_name_columns(headers: list, sample_rows: list[list]) -> tuple[int | None, int | None]:
    normalized_headers = [_normalize_header_text(h) for h in headers]

    # 1) Try header-name based detection first.
    class_col = None
    name_col = None

    for i, h in enumerate(normalized_headers):
        if not h:
            continue
        if name_col is None and ("姓名" in h or h in {"名字", "学生"}):
            name_col = i
        if class_col is None and ("班级" in h or h in {"班别", "班", "班名"}):
            class_col = i

    # 2) If missing, infer from sample row values.
    name_pattern = re.compile(r"^[\u4e00-\u9fa5·]{2,6}$")

    if class_col is None or name_col is None:
        col_count = max((len(r) for r in sample_rows), default=0)
        class_counts = [0] * col_count
        name_counts = [0] * col_count

        for row in sample_rows:
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                text = _normalize_cell_text(cell)
                if not text:
                    continue
                is_class = _looks_like_class_value(text)
                if is_class:
                    class_counts[i] += 1
                # Class-like values such as "三班" should not be counted as names.
                if not is_class and name_pattern.match(text):
                    name_counts[i] += 1

        if class_col is None and class_counts:
            best = max(range(len(class_counts)), key=lambda i: class_counts[i])
            if class_counts[best] > 0:
                class_col = best

        if name_col is None and name_counts:
            best = max(range(len(name_counts)), key=lambda i: name_counts[i])
            if name_counts[best] > 0:
                name_col = best

    # Avoid selecting the same column for both class and name.
    if class_col is not None and name_col == class_col and name_counts:
        ranked_name_cols = sorted(
            range(len(name_counts)),
            key=lambda idx: name_counts[idx],
            reverse=True,
        )
        for idx in ranked_name_cols:
            if idx != class_col and name_counts[idx] > 0:
                name_col = idx
                break

    # 3) Conservative fallback.
    inferred_col_count = max(len(headers), max((len(r) for r in sample_rows), default=0))
    if (class_col is None or name_col is None) and inferred_col_count >= 2:
        if class_col is None:
            class_col = 0 if name_col != 0 else 1
        if name_col is None:
            name_col = 1 if class_col != 1 else 0

    return class_col, name_col


def _parse_exam_grades(exam_grade_value: str | None) -> set[str]:
    return set(parse_grade_tokens(exam_grade_value))


def _canonical_header_name(normalized_header: str) -> str | None:
    for canonical, aliases in HEADER_ALIAS_GROUPS.items():
        if normalized_header in aliases:
            return canonical
    return None


def _build_header_index(headers: list) -> dict[str, int]:
    index_map: dict[str, int] = {}
    for i, header in enumerate(headers):
        canonical = _canonical_header_name(_normalize_header_text(header))
        if canonical and canonical not in index_map:
            index_map[canonical] = i
    return index_map


def _detect_header_row_from_preview(preview_rows: list[list]) -> int | None:
    """Return 1-based row index of the best-matching header row in preview rows."""
    best_row = None
    best_score = -1
    for i, row in enumerate(preview_rows, start=1):
        header_index = _build_header_index(row)
        score = 0
        if "学号" in header_index:
            score += 4
        if "姓名" in header_index:
            score += 2
        if "班级" in header_index:
            score += 1
        if "班级排名" in header_index:
            score += 1
        if "年级排名" in header_index:
            score += 1
        if score > best_score:
            best_score = score
            best_row = i

    if best_score <= 0:
        return None
    return best_row


def _resolve_subject_columns(
    headers: list,
    subject_map: dict[str, Subject],
    sub_headers: list | None = None,
) -> list[tuple[int, Subject]]:
    alias_lookup: dict[str, Subject] = {}

    # Direct subject name mapping from database
    for subject_name, subject in subject_map.items():
        alias_lookup[_normalize_header_text(subject_name)] = subject

    # Common aliases mapping
    for canonical_name, aliases in SUBJECT_ALIAS_GROUPS.items():
        target_subject = subject_map.get(canonical_name)
        if target_subject is None:
            for alias in aliases:
                target_subject = subject_map.get(alias)
                if target_subject is not None:
                    break
        if target_subject is None:
            continue
        for alias in aliases:
            alias_lookup[_normalize_header_text(alias)] = target_subject

    sub_header_markers = {"分数", "客观", "主观", "总名", "校名", "班名", "成绩"}
    use_sub_header_filter = False
    if sub_headers is not None:
        normalized_sub_headers = [_normalize_header_text(h) for h in sub_headers if h is not None]
        use_sub_header_filter = any(h in sub_header_markers for h in normalized_sub_headers)

    subject_cols: list[tuple[int, Subject]] = []
    for i, header in enumerate(headers):
        normalized = _normalize_header_text(header)
        if not normalized:
            continue
        if _canonical_header_name(normalized):
            continue
        if use_sub_header_filter and sub_headers is not None and i < len(sub_headers):
            sub = _normalize_header_text(sub_headers[i])
            if sub and sub not in {"分数", "成绩"}:
                continue
        subject = alias_lookup.get(normalized)
        if subject is not None:
            subject_cols.append((i, subject))

    return subject_cols


def _get_exam_scoped_accessible_class_ids(
    current_user: User,
    db: Session,
    exam: Exam | None,
) -> list[int] | None:
    if current_user.role == "student":
        return None

    base_accessible = get_accessible_class_ids(current_user, db)
    if current_user.role != "teacher" or exam is None:
        return base_accessible
    if current_user.school_id is None:
        return base_accessible

    exam_grades = _parse_exam_grades(exam.grade)
    if not exam_grades:
        return base_accessible

    class_ids = [
        row[0]
        for row in db.query(Class.id)
        .filter(Class.school_id == current_user.school_id, Class.grade.in_(exam_grades))
        .all()
    ]
    return class_ids


def _resolve_student_scope_id(current_user: User, requested_student_id: int | None) -> int | None:
    if current_user.role != "student":
        return requested_student_id
    if current_user.student_id is None:
        raise ValueError("student account is not bound to a student profile")
    return current_user.student_id


def _sort_subjects(subjects: list) -> list:
    """Sort subjects by predefined display order, unknown subjects go to end."""
    order_map = {name: i for i, name in enumerate(SUBJECT_DISPLAY_ORDER)}
    return sorted(subjects, key=lambda s: order_map.get(s.name, 999))


def _get_student_grade(db: Session, student: Student) -> str:
    class_row = db.query(Class).filter(Class.id == student.class_id).first()
    return str(class_row.grade or "").strip() if class_row else ""


def _get_exam_subjects_for_student(
    db: Session,
    exam: Exam,
    student: Student,
    school_id: int | None,
) -> tuple[str, list[Subject]]:
    student_grade = _get_student_grade(db, student)
    subjects = load_exam_subjects_for_grade(db, exam.id, student_grade, school_id)
    return student_grade, _sort_subjects(subjects)


def _validate_exam_subject_scope(
    db: Session,
    exam: Exam,
    student: Student,
    subject_id: int,
    school_id: int | None,
) -> Subject:
    student_grade, subjects = _get_exam_subjects_for_student(db, exam, student, school_id)
    if not student_grade or not subjects:
        raise HTTPException(status_code=400, detail="该学生所在年级未配置本场考试科目")
    subject_map = {int(subject.id): subject for subject in subjects}
    subject = subject_map.get(subject_id)
    if subject is None:
        raise HTTPException(status_code=400, detail="该科目未配置给该学生所在年级的本场考试")
    return subject


def _get_previous_exam(db: Session, current_exam: Exam):
    """Find the previous exam (same grade, earlier date, closest to current)."""
    return (
        db.query(Exam)
        .filter(
            Exam.grade == current_exam.grade,
            Exam.exam_date < current_exam.exam_date,
        )
        .order_by(Exam.exam_date.desc())
        .first()
    )


def _rank_change_str(current_rank, prev_rank):
    if prev_rank is None or current_rank is None:
        return "-"
    diff = prev_rank - current_rank  # positive means improved (rank number decreased)
    if diff > 0:
        return f"↑{diff}"
    elif diff < 0:
        return f"↓{abs(diff)}"
    else:
        return "-"


@router.get("", response_model=ScorePaginatedResponse)
def list_scores(
    exam_id: int,
    class_id: int | None = None,
    student_id: int | None = None,
    student_no: str | None = Query(None, description="学号模糊搜索"),
    student_name: str | None = Query(None, description="姓名模糊搜索"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    # Permission check
    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)

    try:
        resolved_student_id = _resolve_student_scope_id(current_user, student_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if current_user.role == "student":
        bound_student = db.query(Student).filter(Student.id == resolved_student_id).first()
        if bound_student is None:
            raise HTTPException(status_code=400, detail="bound student profile does not exist")

    # Get all students with scores in this exam
    score_query = db.query(Score.student_id).filter(Score.exam_id == exam_id).distinct()
    student_ids_with_scores = [r[0] for r in score_query.all()]

    if not student_ids_with_scores:
        return ScorePaginatedResponse(items=[], total=0, page=page, page_size=page_size)

    # Build student query with filters
    student_query = db.query(Student).filter(Student.id.in_(student_ids_with_scores))
    if class_id:
        if accessible is not None and class_id not in accessible:
            raise HTTPException(status_code=403, detail="无权访问该班级")
        student_query = student_query.filter(Student.class_id == class_id)
    elif accessible is not None:
        student_query = student_query.filter(Student.class_id.in_(accessible))
    if resolved_student_id is not None:
        student_query = student_query.filter(Student.id == resolved_student_id)
    if student_no:
        student_query = student_query.filter(Student.student_no.contains(student_no))
    if student_name:
        student_query = student_query.filter(Student.name.contains(student_name))

    total = student_query.count()

    # Get all matching student IDs for sorting by total_score
    all_matching_student_ids = [s.id for s in student_query.all()]

    if not all_matching_student_ids:
        return ScorePaginatedResponse(items=[], total=total, page=page, page_size=page_size)

    # Get ranks for sorting
    ranks_for_sort = {
        tr.student_id: tr.total_score
        for tr in db.query(TotalRank).filter(
            TotalRank.exam_id == exam_id,
            TotalRank.student_id.in_(all_matching_student_ids),
        ).all()
    }

    # Sort student IDs by total_score DESC
    sorted_student_ids = sorted(
        all_matching_student_ids,
        key=lambda sid: ranks_for_sort.get(sid, 0),
        reverse=True
    )

    # Paginate
    start = (page - 1) * page_size
    end = start + page_size
    paginated_student_ids = sorted_student_ids[start:end]

    if not paginated_student_ids:
        return ScorePaginatedResponse(items=[], total=total, page=page, page_size=page_size)

    # Get students in order
    students_map = {s.id: s for s in db.query(Student).filter(Student.id.in_(paginated_student_ids)).all()}
    students = [students_map[sid] for sid in paginated_student_ids if sid in students_map]

    student_ids = [s.id for s in students]
    class_rows = db.query(Class).all()
    class_map = {c.id: c.name for c in class_rows}
    class_school_map = {c.id: c.school_id for c in class_rows}
    subject_map = {s.id: s.name for s in db.query(Subject).all()}

    # Get scores for these students in this exam
    scores = db.query(Score).filter(
        Score.exam_id == exam_id,
        Score.student_id.in_(student_ids),
    ).all()

    # Group scores by student
    student_scores = {}
    for sc in scores:
        if sc.student_id not in student_scores:
            student_scores[sc.student_id] = {}
        student_scores[sc.student_id][subject_map.get(sc.subject_id, str(sc.subject_id))] = sc.score

    # Get current ranks
    current_ranks = {
        tr.student_id: tr
        for tr in db.query(TotalRank).filter(
            TotalRank.exam_id == exam_id,
            TotalRank.student_id.in_(student_ids),
        ).all()
    }

    # Get previous exam ranks
    prev_exam = _get_previous_exam(db, exam)
    prev_ranks = {}
    if prev_exam:
        prev_ranks = {
            tr.student_id: tr
            for tr in db.query(TotalRank).filter(
                TotalRank.exam_id == prev_exam.id,
                TotalRank.student_id.in_(student_ids),
            ).all()
        }

    # Build response
    full_score_config_cache: dict[int | None, dict[str, float]] = {}
    items = []
    for s in students:
        subj_scores = student_scores.get(s.id, {})
        curr_rank = current_ranks.get(s.id)
        prev_rank = prev_ranks.get(s.id)
        school_id_for_full_score = class_school_map.get(s.class_id) or get_user_school_id(current_user)
        if school_id_for_full_score not in full_score_config_cache:
            full_score_config_cache[school_id_for_full_score] = _load_subject_full_score_config(
                db,
                school_id_for_full_score,
            )
        full_score_config = full_score_config_cache[school_id_for_full_score]
        subject_full_scores = {
            subject_name: _subject_full_score_for_three_rates(
                subject_name,
                full_score_config=full_score_config,
            )
            for subject_name in subj_scores
        }

        items.append(ScoreItemResponse(
            student_id=s.id,
            student_no=s.student_no,
            student_name=s.name,
            class_name=class_map.get(s.class_id, ""),
            subjects=subj_scores,
            subject_full_scores=subject_full_scores,
            total_score=curr_rank.total_score if curr_rank else sum(subj_scores.values()),
            rank_class=curr_rank.rank_class if curr_rank else None,
            rank_grade=curr_rank.rank_grade if curr_rank else None,
            prev_total_score=prev_rank.total_score if prev_rank else None,
            prev_rank_class=prev_rank.rank_class if prev_rank else None,
            prev_rank_grade=prev_rank.rank_grade if prev_rank else None,
            rank_class_change=_rank_change_str(
                curr_rank.rank_class if curr_rank else None,
                prev_rank.rank_class if prev_rank else None,
            ),
            rank_grade_change=_rank_change_str(
                curr_rank.rank_grade if curr_rank else None,
                prev_rank.rank_grade if prev_rank else None,
            ),
        ))

    return ScorePaginatedResponse(items=items, total=total, page=page, page_size=page_size)


@router.get("/entry-subjects")
def get_entry_subjects(
    exam_id: int,
    student_id: int,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该学生")

    school_id = get_user_school_id(current_user)
    student_grade, subjects = _get_exam_subjects_for_student(db, exam, student, school_id)
    return {
        "grade": student_grade,
        "subjects": [{"id": int(subject.id), "name": subject.name} for subject in subjects],
    }


@router.post("", status_code=status.HTTP_201_CREATED)
def create_score(
    req: ScoreCreate,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    exam = db.query(Exam).filter(Exam.id == req.exam_id).first()
    if not exam:
        raise HTTPException(status_code=400, detail="考试不存在")

    student = db.query(Student).filter(Student.id == req.student_id).first()
    if not student:
        raise HTTPException(status_code=400, detail="学生不存在")

    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该学生")

    school_id = get_user_school_id(current_user)
    _validate_exam_subject_scope(db, exam, student, req.subject_id, school_id)

    existing = db.query(Score).filter(
        Score.student_id == req.student_id,
        Score.exam_id == req.exam_id,
        Score.subject_id == req.subject_id,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail="该学生此科目成绩已录入")

    score = Score(
        student_id=req.student_id,
        exam_id=req.exam_id,
        subject_id=req.subject_id,
        score=req.score,
    )
    db.add(score)
    db.flush()
    recalculate_ranks(db, req.exam_id)
    db.commit()
    return {"id": score.id, "message": "成绩录入成功"}


@router.put("/upsert")
def upsert_score(
    req: ScoreCreate,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    """Create or update a score record."""
    exam = db.query(Exam).filter(Exam.id == req.exam_id).first()
    if not exam:
        raise HTTPException(status_code=400, detail="考试不存在")

    student = db.query(Student).filter(Student.id == req.student_id).first()
    if not student:
        raise HTTPException(status_code=400, detail="学生不存在")

    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该学生")

    school_id = get_user_school_id(current_user)
    _validate_exam_subject_scope(db, exam, student, req.subject_id, school_id)

    existing = db.query(Score).filter(
        Score.student_id == req.student_id,
        Score.exam_id == req.exam_id,
        Score.subject_id == req.subject_id,
    ).first()

    if existing:
        existing.score = req.score
    else:
        db.add(
            Score(
                student_id=req.student_id,
                exam_id=req.exam_id,
                subject_id=req.subject_id,
                score=req.score,
            )
        )

    db.flush()
    recalculate_ranks(db, req.exam_id)
    db.commit()
    return {"message": "成绩保存成功"}


@router.put("/{score_id}")
def update_score(
    score_id: int,
    req: ScoreUpdate,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    score = db.query(Score).filter(Score.id == score_id).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩记录不存在")
    student = db.query(Student).filter(Student.id == score.student_id).first()
    exam = db.query(Exam).filter(Exam.id == score.exam_id).first()
    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    if accessible is not None and student and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该成绩")

    score.score = req.score
    db.flush()
    recalculate_ranks(db, score.exam_id)
    db.commit()
    return {"message": "成绩修改成功"}


@router.delete("/by-student", status_code=status.HTTP_204_NO_CONTENT)
def delete_scores_by_student(
    exam_id: int = Query(...),
    student_id: int = Query(...),
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    """Delete all scores for a specific student in a specific exam."""
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="学生不存在")

    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    if accessible is not None and student.class_id not in accessible:
        raise HTTPException(status_code=403, detail="无权操作该学生")

    scores = db.query(Score).filter(
        Score.exam_id == exam_id,
        Score.student_id == student_id,
    ).all()

    for score in scores:
        db.delete(score)

    # Also delete TotalRank
    total_rank = db.query(TotalRank).filter(
        TotalRank.exam_id == exam_id,
        TotalRank.student_id == student_id,
    ).first()
    if total_rank:
        db.delete(total_rank)

    db.flush()
    recalculate_ranks(db, exam_id)
    db.commit()


@router.delete("/{score_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_score(
    score_id: int,
    _: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    score = db.query(Score).filter(Score.id == score_id).first()
    if not score:
        raise HTTPException(status_code=404, detail="成绩记录不存在")
    exam_id = score.exam_id
    db.delete(score)
    db.flush()
    recalculate_ranks(db, exam_id)
    db.commit()


@router.post("/batch-by-students/delete")
def batch_delete_scores_by_students(
    req: BatchDeleteScoresRequest,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    """Batch delete scores for multiple students in a specific exam."""
    if not req.student_ids:
        raise HTTPException(status_code=400, detail="请提供 student_ids")

    exam = db.query(Exam).filter(Exam.id == req.exam_id).first()
    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    students = db.query(Student).filter(Student.id.in_(req.student_ids)).all()

    deleted_count = 0
    for student in students:
        if accessible is not None and student.class_id not in accessible:
            continue

        scores = db.query(Score).filter(
            Score.exam_id == req.exam_id,
            Score.student_id == student.id,
        ).all()

        for score in scores:
            db.delete(score)
            deleted_count += 1

        # Also delete TotalRank
        total_rank = db.query(TotalRank).filter(
            TotalRank.exam_id == req.exam_id,
            TotalRank.student_id == student.id,
        ).first()
        if total_rank:
            db.delete(total_rank)

    db.flush()
    recalculate_ranks(db, req.exam_id)
    db.commit()

    return {"deleted_count": deleted_count}


@router.get("/template")
def download_score_template(
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    school_id = get_user_school_id(current_user)
    subject_query = db.query(Subject)
    if school_id is not None:
        subject_query = subject_query.filter(Subject.school_id == school_id)
    subjects = _sort_subjects(subject_query.all())
    wb = Workbook()
    ws = wb.active
    ws.title = "成绩导入模板"
    headers = ["学号", "班级", "姓名"]
    for subj in subjects:
        header = subj.name
        if subj.name in REQUIRED_SUBJECTS:
            header += "(必填)"
        headers.append(header)
    ws.append(headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=score_template.xlsx"},
    )


@router.get("/export")
def export_scores(
    exam_id: int,
    class_id: int | None = None,
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=404, detail="考试不存在")

    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    school_id = get_user_school_id(current_user)
    class_rows = db.query(Class).all()
    class_map = {c.id: c.name for c in class_rows}
    class_grade_map = {c.id: c.grade for c in class_rows}

    score_query = db.query(Score).filter(Score.exam_id == exam_id)
    student_ids = [r[0] for r in score_query.with_entities(Score.student_id).distinct().all()]

    student_query = db.query(Student).filter(Student.id.in_(student_ids))
    if class_id:
        student_query = student_query.filter(Student.class_id == class_id)
    elif accessible is not None:
        student_query = student_query.filter(Student.class_id.in_(accessible))

    students = student_query.all()
    sid_list = [s.id for s in students]

    scores = db.query(Score).filter(Score.exam_id == exam_id, Score.student_id.in_(sid_list)).all()
    ranks = {
        tr.student_id: tr
        for tr in db.query(TotalRank).filter(TotalRank.exam_id == exam_id, TotalRank.student_id.in_(sid_list)).all()
    }

    student_scores: dict[int, dict[int, float]] = {}
    for sc in scores:
        student_scores.setdefault(sc.student_id, {})[sc.subject_id] = sc.score

    wb = Workbook()
    wb.remove(wb.active)

    grade_order = parse_grade_tokens(exam.grade)
    for grade in grade_order:
        grade_subjects = load_exam_subjects_for_grade(db, exam.id, grade, school_id)
        if not grade_subjects:
            continue

        grade_students = [student for student in students if class_grade_map.get(student.class_id) == grade]
        ws = wb.create_sheet(title=grade)
        headers = ["学号", "姓名", "班级"] + [subject.name for subject in grade_subjects] + ["总分", "班级排名", "年级排名"]
        ws.append(headers)

        for student in grade_students:
            row = [student.student_no, student.name, class_map.get(student.class_id, "")]
            score_map = student_scores.get(student.id, {})
            for subject in grade_subjects:
                row.append(score_map.get(subject.id, ""))
            rank = ranks.get(student.id)
            total_score = rank.total_score if rank else sum(score_map.get(subject.id, 0) for subject in grade_subjects)
            row.append(total_score if total_score else "")
            row.append(rank.rank_class if rank else "")
            row.append(rank.rank_grade if rank else "")
            ws.append(row)

    if not wb.sheetnames:
        wb.create_sheet(title="成绩")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=scores_{exam_id}.xlsx"},
    )


@router.post("/import")
def import_scores(
    exam_id: int = Query(..., description="考试ID"),
    grade: str = Query(..., description="导入年级"),
    file: UploadFile = File(...),
    current_user: User = Depends(require_teacher_or_admin),
    db: Session = Depends(get_db),
):

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件")

    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(status_code=400, detail="考试不存在")

    selected_grade = grade.strip()
    if not selected_grade:
        raise HTTPException(status_code=400, detail="请先选择导入年级")

    exam_grades = _parse_exam_grades(exam.grade)
    if exam_grades and selected_grade not in exam_grades:
        raise HTTPException(status_code=400, detail="导入年级与考试参与年级不一致")

    accessible = _get_exam_scoped_accessible_class_ids(current_user, db, exam)
    school_id = get_user_school_id(current_user)

    # Build lookup maps - filter by school
    class_query = db.query(Class)
    if school_id is not None:
        class_query = class_query.filter(Class.school_id == school_id)
    classes = class_query.all()
    class_grade_map = {c.id: c.grade for c in classes}
    grade_class_name_map = {c.id: c.name for c in classes if c.grade == selected_grade}

    student_query = db.query(Student)
    if school_id is not None:
        student_query = (
            student_query.join(Class, Class.id == Student.class_id)
            .filter(Class.school_id == school_id)
        )
    students = student_query.all()
    student_map = {s.student_no: s for s in students}
    students_by_class_and_name: dict[tuple[int, str], list[Student]] = {}
    for s in students:
        if class_grade_map.get(s.class_id) != selected_grade:
            continue
        key = (s.class_id, _normalize_header_text(s.name))
        students_by_class_and_name.setdefault(key, []).append(s)

    subjects = load_exam_subjects_for_grade(db, exam.id, selected_grade, school_id)
    if not subjects:
        raise HTTPException(status_code=400, detail="????????????")
    subject_map = {s.name: s for s in subjects}

    wb = load_workbook(file.file)
    ws = wb.active
    preview_row_count = min(10, ws.max_row)
    preview_rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, preview_row_count + 1)
    ]
    header_row = _detect_header_row_from_preview(preview_rows)
    if header_row is None:
        raise HTTPException(status_code=400, detail="Excel 未识别到表头，请确认包含学号/考号列")

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    sub_headers = []
    if header_row < ws.max_row:
        sub_headers = [ws.cell(header_row + 1, c).value for c in range(1, ws.max_column + 1)]
    header_index = _build_header_index(headers)

    student_no_col = header_index.get("学号")
    class_col = header_index.get("班级")
    name_col = header_index.get("姓名")

    sample_rows = list(ws.iter_rows(min_row=header_row + 1, max_row=min(ws.max_row, header_row + 10), values_only=True))
    if student_no_col is None and (class_col is None or name_col is None):
        inferred_class_col, inferred_name_col = _infer_class_name_columns(headers, sample_rows)
        if class_col is None:
            class_col = inferred_class_col
        if name_col is None:
            name_col = inferred_name_col

    if student_no_col is None and (class_col is None or name_col is None):
        raise HTTPException(status_code=400, detail="Excel 缺少学号列，且无法通过班级+姓名识别学生")

    # Detect rank columns (班级排名/年级排名)
    rank_class_col = header_index.get("班级排名")
    rank_grade_col = header_index.get("年级排名")

    subject_cols = _resolve_subject_columns(headers, subject_map, sub_headers=sub_headers)
    if not subject_cols:
        raise HTTPException(status_code=400, detail="Excel 未识别到任何科目列")

    rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
    success_count = 0
    errors = []
    # Track students with explicit ranks from Excel
    students_with_explicit_ranks = {}

    for idx, row in enumerate(rows, start=header_row + 1):
        try:
            if not row:
                continue
            row_has_data = any(_normalize_cell_text(cell) != "" for cell in row)
            if not row_has_data:
                continue

            student = None
            student_no = None
            pending_student = None

            student_no_value = (
                row[student_no_col]
                if student_no_col is not None and student_no_col < len(row)
                else None
            )
            student_no_text = _normalize_cell_text(student_no_value)
            has_student_no = bool(student_no_text)

            if has_student_no:
                student_no = student_no_text
                if student_no not in student_map:
                    if class_col is None or name_col is None:
                        errors.append({"row": idx, "error": f"学号 '{student_no}' 不存在，且缺少班级或姓名，无法自动创建学生"})
                        continue

                    class_name = (
                        _normalize_cell_text(row[class_col])
                        if class_col < len(row) and row[class_col] is not None
                        else ""
                    )
                    student_name = (
                        _normalize_cell_text(row[name_col])
                        if name_col < len(row) and row[name_col] is not None
                        else ""
                    )

                    if class_name and student_name and not _looks_like_class_value(class_name) and _looks_like_class_value(student_name):
                        class_name, student_name = student_name, class_name

                    if not class_name or not student_name:
                        errors.append({"row": idx, "error": f"学号 '{student_no}' 不存在，且缺少班级或姓名，无法自动创建学生"})
                        continue

                    class_ids = _resolve_class_ids_by_name(class_name, grade_class_name_map)
                    if not class_ids:
                        errors.append({"row": idx, "error": f"班级 '{class_name}' 在年级 '{selected_grade}' 中不存在"})
                        continue
                    if len(class_ids) > 1:
                        errors.append({"row": idx, "error": f"班级 '{class_name}' 匹配到多个班级，请先维护学生档案"})
                        continue
                    if accessible is not None and class_ids[0] not in accessible:
                        errors.append({"row": idx, "error": f"学号 '{student_no}' 无权导入"})
                        continue

                    pending_student = Student(
                        student_no=student_no,
                        name=student_name,
                        gender="U",
                        class_id=class_ids[0],
                    )
                    student = pending_student
                else:
                    student = student_map[student_no]
            else:
                if class_col is None or name_col is None:
                    errors.append({"row": idx, "error": "缺少学号且无法通过班级+姓名识别"})
                    continue
                class_name = (
                    _normalize_cell_text(row[class_col])
                    if class_col < len(row) and row[class_col] is not None
                    else ""
                )
                student_name = (
                    _normalize_cell_text(row[name_col])
                    if name_col < len(row) and row[name_col] is not None
                    else ""
                )

                if class_name and student_name and not _looks_like_class_value(class_name) and _looks_like_class_value(student_name):
                    class_name, student_name = student_name, class_name

                if not class_name and not student_name:
                    # e.g. merged-subheader row in multi-level header excel
                    continue
                if not class_name or not student_name:
                    errors.append({"row": idx, "error": "缺少班级或姓名，无法识别学生"})
                    continue

                class_ids = _resolve_class_ids_by_name(class_name, grade_class_name_map)
                if not class_ids:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 在年级 '{selected_grade}' 中不存在"})
                    continue
                if len(class_ids) > 1:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 匹配到多个班级，请改为学号导入"})
                    continue

                key = (class_ids[0], _normalize_header_text(student_name))
                candidates = students_by_class_and_name.get(key, [])
                if len(candidates) == 0:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 下未找到姓名 '{student_name}' 的学生"})
                    continue
                if len(candidates) > 1:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 下姓名 '{student_name}' 重复，请改为学号导入"})
                    continue
                student = candidates[0]
                student_no = student.student_no

            student_grade = class_grade_map.get(student.class_id)
            if student_grade != selected_grade:
                errors.append({"row": idx, "error": f"学生 '{student_no}' 不属于所选年级 '{selected_grade}'"})
                continue
            if accessible is not None and student.class_id not in accessible:
                errors.append({"row": idx, "error": f"学号 '{student_no}' 无权导入"})
                continue

            # Check required subjects have values
            missing_required = []
            required_subject_names = {
                subject.name for _, subject in subject_cols if subject.name in REQUIRED_SUBJECTS
            }
            for subject_name in required_subject_names:
                has_value = any(
                    subject.name == subject_name
                    and col_idx < len(row)
                    and row[col_idx] is not None
                    and str(row[col_idx]).strip() != ""
                    for col_idx, subject in subject_cols
                )
                if not has_value:
                    missing_required.append(subject_name)
            missing_required_set = set(missing_required)

            if pending_student is not None:
                db.add(pending_student)
                db.flush()
                student_map[student_no] = pending_student
                key = (pending_student.class_id, _normalize_header_text(pending_student.name))
                students_by_class_and_name.setdefault(key, []).append(pending_student)

            for col_idx, subject in subject_cols:
                has_score_value = col_idx < len(row) and row[col_idx] is not None and str(row[col_idx]).strip() != ""
                if has_score_value or subject.name in missing_required_set:
                    try:
                        score_val = float(row[col_idx]) if has_score_value else 0
                    except (ValueError, TypeError):
                        errors.append({"row": idx, "error": f"科目 '{subject.name}' 分数格式错误"})
                        continue

                    existing = db.query(Score).filter(
                        Score.student_id == student.id,
                        Score.exam_id == exam.id,
                        Score.subject_id == subject.id,
                    ).first()
                    if existing:
                        existing.score = score_val
                    else:
                        db.add(Score(
                            student_id=student.id,
                            exam_id=exam.id,
                            subject_id=subject.id,
                            score=score_val,
                        ))

            # Check for explicit rank values
            explicit_rank_class = None
            explicit_rank_grade = None
            if rank_class_col is not None and rank_class_col < len(row) and row[rank_class_col]:
                try:
                    explicit_rank_class = int(row[rank_class_col])
                except (ValueError, TypeError):
                    pass
            if rank_grade_col is not None and rank_grade_col < len(row) and row[rank_grade_col]:
                try:
                    explicit_rank_grade = int(row[rank_grade_col])
                except (ValueError, TypeError):
                    pass

            if explicit_rank_class is not None or explicit_rank_grade is not None:
                students_with_explicit_ranks[student.id] = {
                    "rank_class": explicit_rank_class,
                    "rank_grade": explicit_rank_grade,
                }

            success_count += 1
        except Exception as e:
            errors.append({"row": idx, "error": str(e)})

    db.flush()

    # Use explicit ranks only when all imported rows provide both class and grade ranks.
    if _should_use_explicit_ranks(students_with_explicit_ranks, success_count):
        # Use explicit ranks - update TotalRank directly
        for student_id, ranks in students_with_explicit_ranks.items():
            # Calculate total score for this student
            scores = db.query(Score).filter(
                Score.exam_id == exam_id,
                Score.student_id == student_id,
            ).all()
            total_score = sum(s.score for s in scores)

            existing_rank = db.query(TotalRank).filter(
                TotalRank.exam_id == exam_id,
                TotalRank.student_id == student_id,
            ).first()
            if existing_rank:
                existing_rank.total_score = total_score
                existing_rank.rank_class = ranks["rank_class"]
                existing_rank.rank_grade = ranks["rank_grade"]
            else:
                db.add(TotalRank(
                    exam_id=exam_id,
                    student_id=student_id,
                    total_score=total_score,
                    rank_class=ranks["rank_class"],
                    rank_grade=ranks["rank_grade"],
                ))
    else:
        # Recalculate ranks as before
        recalculate_ranks(db, exam_id)

    db.commit()

    return {"success_count": success_count, "error_count": len(errors), "errors": errors}
