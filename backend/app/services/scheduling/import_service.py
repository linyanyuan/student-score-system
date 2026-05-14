from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.class_ import Class
from app.models.schedule_draft import ScheduleDraft
from app.models.schedule_draft_item import ScheduleDraftItem
from app.models.schedule_import import ScheduleImport
from app.models.schedule_import_item import ScheduleImportItem
from app.models.schedule_period import SchedulePeriod
from app.models.subject import Subject
from app.models.teacher_class_subject import TeacherClassSubject
from app.models.user import User

WEEKDAY_ALIASES = {
    "周一": 1,
    "星期一": 1,
    "礼拜一": 1,
    "一": 1,
    "周二": 2,
    "星期二": 2,
    "礼拜二": 2,
    "二": 2,
    "周三": 3,
    "星期三": 3,
    "礼拜三": 3,
    "三": 3,
    "周四": 4,
    "星期四": 4,
    "礼拜四": 4,
    "四": 4,
    "周五": 5,
    "星期五": 5,
    "礼拜五": 5,
    "五": 5,
}

EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SUBJECT_ALIAS_TARGETS = {
    "语": ["语文"],
    "数": ["数学"],
    "英": ["英语"],
    "物": ["物理"],
    "化": ["化学"],
    "生": ["生物"],
    "政": ["政治", "道法", "道德与法治"],
    "道": ["道法", "道德与法治"],
    "地": ["地理"],
    "历": ["历史"],
    "史": ["历史"],
    "体": ["体育"],
    "音": ["音乐"],
    "美": ["美术"],
    "微": ["微机", "信息技术"],
    "信": ["信息技术", "微机"],
    "劳": ["劳技", "劳动"],
    "班": ["班会"],
}
CHINESE_NUMBERS = {
    "零": 0,
    "一": 1,
    "二": 2,
    "两": 2,
    "三": 3,
    "四": 4,
    "五": 5,
    "六": 6,
    "七": 7,
    "八": 8,
    "九": 9,
    "十": 10,
}


@dataclass(frozen=True)
class ParsedSlot:
    class_id: int
    weekday: int
    period_id: int
    subject_text: str
    teacher_text: str | None = None


def _json_dumps(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False)


def _json_loads(payload: str | None, fallback: Any) -> Any:
    if not payload:
        return fallback
    try:
        return json.loads(payload)
    except json.JSONDecodeError:
        return fallback


def _file_extension(filename: str | None) -> str:
    if not filename or "." not in filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower()


def detect_source_type(file: UploadFile) -> str:
    extension = _file_extension(file.filename)
    if extension in EXCEL_EXTENSIONS:
        return "excel"
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="仅支持 Excel 课表导入")


def _normalize_text(value: Any) -> str:
    return str(value or "").strip()


def _compact_name(value: str) -> str:
    return re.sub(r"\s+", "", value).strip()


def _parse_weekday(value: Any) -> int | None:
    text = _compact_name(_normalize_text(value))
    if text in WEEKDAY_ALIASES:
        return WEEKDAY_ALIASES[text]
    match = re.search(r"([1-5])", text)
    if match:
        return int(match.group(1))
    return None


def _parse_chinese_number(text: str) -> int | None:
    if not text:
        return None
    if text in CHINESE_NUMBERS:
        return CHINESE_NUMBERS[text]
    if text.startswith("十") and len(text) == 2 and text[1] in CHINESE_NUMBERS:
        return 10 + CHINESE_NUMBERS[text[1]]
    if text.endswith("十") and len(text) == 2 and text[0] in CHINESE_NUMBERS:
        return CHINESE_NUMBERS[text[0]] * 10
    if "十" in text and len(text) == 3:
        left, right = text.split("十", 1)
        if left in CHINESE_NUMBERS and right in CHINESE_NUMBERS:
            return CHINESE_NUMBERS[left] * 10 + CHINESE_NUMBERS[right]
    return None


def _parse_period_ordinal(text: str) -> int | None:
    match = re.search(r"(\d+)", text)
    if match:
        return int(match.group(1))
    match = re.search(r"第?([一二两三四五六七八九十]{1,3})节?", text)
    if match:
        return _parse_chinese_number(match.group(1))
    return None


def _parse_period(value: Any, periods: list[SchedulePeriod]) -> int | None:
    text = _compact_name(_normalize_text(value))
    if not text:
        return None
    by_name = {_compact_name(item.name): item.id for item in periods}
    if text in by_name:
        return by_name[text]
    number = _parse_period_ordinal(text)
    if number is not None:
        if 1 <= number <= len(periods):
            return periods[number - 1].id
        for period in periods:
            if period.sort_order == number:
                return period.id
    return None


def _normalize_class_sheet_name(value: Any) -> str:
    text = _compact_name(_normalize_text(value))
    text = text.replace("年级", "")
    text = text.removesuffix("班")
    result = []
    for char in text:
        if char in CHINESE_NUMBERS and char != "十":
            result.append(str(CHINESE_NUMBERS[char]))
        elif char == "十":
            result.append("10")
        else:
            result.append(char)
    return "".join(result).lower()


def _split_cell_text(value: Any) -> tuple[str, str | None]:
    text = _normalize_text(value)
    if not text:
        return "", None
    parts = [part.strip() for part in re.split(r"[/／\n\r]+", text) if part.strip()]
    if not parts:
        return "", None
    if len(parts) == 1:
        return parts[0], None
    return parts[0], parts[1]


def _subject_map(subjects: list[Subject]) -> dict[str, Subject]:
    result: dict[str, Subject] = {}
    for subject in subjects:
        result[_compact_name(subject.name)] = subject
        result[_compact_name(subject.code)] = subject
    for alias, target_names in SUBJECT_ALIAS_TARGETS.items():
        for target_name in target_names:
            subject = result.get(_compact_name(target_name))
            if subject is not None:
                result.setdefault(alias, subject)
                break
    return result


def _find_subject(subject_text: str, subjects: list[Subject]) -> Subject | None:
    compact = _compact_name(subject_text)
    if not compact:
        return None
    return _subject_map(subjects).get(compact)


def _find_teacher_by_text(db: Session, school_id: int, teacher_text: str | None) -> User | None:
    if not teacher_text:
        return None
    compact = _compact_name(teacher_text)
    compact = compact.strip("()（）").removesuffix("老师")
    teachers = db.query(User).filter(User.school_id == school_id, User.role == "teacher").all()
    for teacher in teachers:
        username = _compact_name(teacher.username).strip("()（）").removesuffix("老师")
        if username == compact:
            return teacher
    return None


def _teacher_candidates(db: Session, school_id: int, class_id: int, subject_id: int) -> list[User]:
    rows = (
        db.query(User)
        .join(TeacherClassSubject, TeacherClassSubject.teacher_id == User.id)
        .filter(
            TeacherClassSubject.school_id == school_id,
            TeacherClassSubject.class_id == class_id,
            TeacherClassSubject.subject_id == subject_id,
            User.role == "teacher",
        )
        .all()
    )
    return rows


def _candidate_payload(candidates: list[User]) -> list[dict[str, Any]]:
    return [{"id": item.id, "username": item.username} for item in candidates]


def build_issue_flags(item: ScheduleImportItem) -> list[str]:
    if item.is_empty:
        return []
    flags: list[str] = []
    if item.subject_id is None:
        flags.append("unrecognized_subject")
    if item.teacher_match_status == "ambiguous":
        flags.append("teacher_ambiguous")
    elif item.subject_id is not None and item.teacher_id is None:
        flags.append("teacher_unmatched")
    return flags


def _apply_item_flags(item: ScheduleImportItem) -> None:
    item.issue_flags = _json_dumps(build_issue_flags(item))


def _teacher_time_conflict_item_ids(rows: list[ScheduleImportItem]) -> set[int]:
    slot_items: dict[tuple[int, int, int], list[ScheduleImportItem]] = {}
    for row in rows:
        if row.is_empty or row.teacher_id is None:
            continue
        slot_items.setdefault((row.teacher_id, row.weekday, row.period_id), []).append(row)

    conflict_ids: set[int] = set()
    for items in slot_items.values():
        if len(items) > 1:
            conflict_ids.update(item.id for item in items if item.id is not None)
    return conflict_ids


def rebuild_import_item_flags(db: Session, schedule_import: ScheduleImport) -> list[ScheduleImportItem]:
    rows = db.query(ScheduleImportItem).filter(ScheduleImportItem.import_id == schedule_import.id).all()
    conflict_ids = _teacher_time_conflict_item_ids(rows)
    for row in rows:
        flags = build_issue_flags(row)
        if row.id in conflict_ids:
            flags.append("teacher_time_conflict")
        row.issue_flags = _json_dumps(flags)
    return rows


def rebuild_import_summary(db: Session, schedule_import: ScheduleImport) -> dict[str, int]:
    rows = db.query(ScheduleImportItem).filter(ScheduleImportItem.import_id == schedule_import.id).all()
    summary = {
        "total_slots": len(rows),
        "recognized_slots": 0,
        "unrecognized_subject_slots": 0,
        "teacher_unmatched_slots": 0,
        "teacher_ambiguous_slots": 0,
        "teacher_time_conflict_slots": 0,
        "manually_fixed_slots": 0,
    }
    for row in rows:
        flags = set(_json_loads(row.issue_flags, []))
        if row.subject_id is not None:
            summary["recognized_slots"] += 1
        if "unrecognized_subject" in flags:
            summary["unrecognized_subject_slots"] += 1
        if "teacher_unmatched" in flags:
            summary["teacher_unmatched_slots"] += 1
        if "teacher_ambiguous" in flags:
            summary["teacher_ambiguous_slots"] += 1
        if "teacher_time_conflict" in flags:
            summary["teacher_time_conflict_slots"] += 1
        if "manual" in _json_loads(row.teacher_candidates, []):
            summary["manually_fixed_slots"] += 1
    schedule_import.summary = _json_dumps(summary)
    return summary


def _load_excel_workbook(file: UploadFile):
    try:
        file.file.seek(0)
        return load_workbook(file.file, data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="无法读取 Excel 课表") from exc


def _find_weekday_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, int]]:
    for row_index, row in enumerate(rows[:10]):
        weekday_columns: dict[int, int] = {}
        for index, value in enumerate(row):
            weekday = _parse_weekday(value)
            if weekday is not None:
                weekday_columns[index] = weekday
        has_period_header = any("节次" in _compact_name(_normalize_text(value)) for value in row)
        if len(weekday_columns) >= 2 or (weekday_columns and has_period_header):
            return row_index, weekday_columns
    return 0, {}


def _parse_sheet_slots(sheet: Any, *, class_id: int, periods: list[SchedulePeriod]) -> list[ParsedSlot]:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    header_index, weekday_columns = _find_weekday_header(rows)

    slots: list[ParsedSlot] = []
    for row in rows[header_index + 1 :]:
        if not row:
            continue
        first_weekday_column = min(weekday_columns.keys()) if weekday_columns else len(row)
        period_id = None
        for value in row[:first_weekday_column]:
            period_id = _parse_period(value, periods)
            if period_id is not None:
                break
        if period_id is None:
            continue
        for column_index, weekday in weekday_columns.items():
            if column_index >= len(row):
                continue
            subject_text, teacher_text = _split_cell_text(row[column_index])
            if not subject_text:
                continue
            slots.append(ParsedSlot(class_id=class_id, weekday=weekday, period_id=period_id, subject_text=subject_text, teacher_text=teacher_text))
    return slots


def _parse_excel_slots(file: UploadFile, *, class_id: int, periods: list[SchedulePeriod]) -> list[ParsedSlot]:
    workbook = _load_excel_workbook(file)
    return _parse_sheet_slots(workbook.active, class_id=class_id, periods=periods)


def _parse_grade_excel_slots(file: UploadFile, *, classes: list[Class], periods: list[SchedulePeriod]) -> tuple[list[ParsedSlot], list[str]]:
    workbook = _load_excel_workbook(file)
    class_by_name = {_normalize_class_sheet_name(item.name): item for item in classes}
    slots: list[ParsedSlot] = []
    unmatched_sheets: list[str] = []
    for sheet in workbook.worksheets:
        class_obj = class_by_name.get(_normalize_class_sheet_name(sheet.title))
        if class_obj is None:
            unmatched_sheets.append(sheet.title)
            continue
        slots.extend(_parse_sheet_slots(sheet, class_id=class_obj.id, periods=periods))
    return slots, unmatched_sheets


def _import_period_query(db: Session, school_id: int | None) -> list[SchedulePeriod]:
    query = db.query(SchedulePeriod).filter(
        SchedulePeriod.is_active == True,  # noqa: E712
        SchedulePeriod.include_in_auto_schedule == True,  # noqa: E712
    )
    if school_id is not None:
        school_periods = (
            query.filter(SchedulePeriod.school_id == school_id)
            .order_by(SchedulePeriod.sort_order, SchedulePeriod.id)
            .all()
        )
        if school_periods:
            return school_periods
    return query.filter(SchedulePeriod.school_id.is_(None)).order_by(SchedulePeriod.sort_order, SchedulePeriod.id).all()


def build_schedule_import_template() -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "课表模板"
    headers = ["节次", "周一", "周二", "周三", "周四", "周五"]
    sheet.append(headers)
    for index in range(1, 8):
        sheet.append([f"第{index}节", "", "", "", "", ""])
    sheet["B2"] = "数学"
    sheet["C2"] = "语文/张老师"
    sheet.freeze_panes = "B2"
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column in range(1, 7):
        sheet.column_dimensions[chr(64 + column)].width = 18
    sheet["H1"] = "填写说明"
    sheet["H2"] = "单元格可只填科目，例如：数学"
    sheet["H3"] = "也可填 科目/教师，例如：语文/张老师"
    sheet["H4"] = "系统会根据班级-科目-教师自动匹配教师"
    sheet.column_dimensions["H"].width = 42
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_import_from_upload(
    db: Session,
    *,
    school_id: int,
    grade: str,
    scope: str,
    class_id: int | None,
    file: UploadFile,
    created_by: int | None,
) -> ScheduleImport:
    source_type = detect_source_type(file)
    if scope not in {"grade", "class"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="invalid import scope")

    target_class: Class | None = None
    target_classes: list[Class] = []
    if scope == "class":
        if class_id is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="class_id is required for class import")
        target_class = db.query(Class).filter(Class.id == class_id, Class.school_id == school_id, Class.grade == grade).first()
        if target_class is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="class not found")
    else:
        target_classes = db.query(Class).filter(Class.school_id == school_id, Class.grade == grade).order_by(Class.id).all()
        if not target_classes:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="class not found")

    schedule_import = ScheduleImport(
        school_id=school_id,
        grade=grade,
        scope=scope,
        class_id=class_id,
        source_type=source_type,
        status="processing",
        message="正在识别课表",
        created_by=created_by,
    )
    db.add(schedule_import)
    db.flush()

    periods = _import_period_query(db, school_id)
    if not periods:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="请先配置参与排课的节次")
    subjects = db.query(Subject).filter(Subject.school_id == school_id).all()

    unmatched_sheets: list[str] = []
    if scope == "grade":
        slots, unmatched_sheets = _parse_grade_excel_slots(file, classes=target_classes, periods=periods)
    else:
        slots = _parse_excel_slots(file, class_id=target_class.id, periods=periods)
    for slot in slots:
        subject = _find_subject(slot.subject_text, subjects)
        teacher: User | None = None
        match_status = "unmatched"
        match_source = "unmatched"
        candidates: list[User] = []
        if subject is not None:
            teacher = _find_teacher_by_text(db, school_id, slot.teacher_text)
            if teacher is not None:
                match_status = "matched"
                match_source = "file_recognition"
            else:
                candidates = _teacher_candidates(db, school_id, slot.class_id, subject.id)
                if len(candidates) == 1:
                    teacher = candidates[0]
                    match_status = "matched"
                    match_source = "teaching_arrangement"
                elif len(candidates) > 1:
                    match_status = "ambiguous"
                    match_source = "teaching_arrangement"
                else:
                    match_status = "unmatched"
                    match_source = "unmatched"

        item = ScheduleImportItem(
            import_id=schedule_import.id,
            class_id=slot.class_id,
            weekday=slot.weekday,
            period_id=slot.period_id,
            subject_id=subject.id if subject else None,
            recognized_subject_name=slot.subject_text,
            teacher_id=teacher.id if teacher else None,
            teacher_match_status=match_status,
            teacher_match_source=match_source,
            teacher_candidates=_json_dumps(_candidate_payload(candidates)),
            confidence=1.0 if subject else 0.0,
            is_empty=False,
        )
        _apply_item_flags(item)
        db.add(item)

    schedule_import.status = "needs_review"
    schedule_import.message = "已生成识别结果，请核对后生成待确认草案"
    if unmatched_sheets:
        preview = "、".join(unmatched_sheets[:5])
        suffix = "等" if len(unmatched_sheets) > 5 else ""
        schedule_import.message = f"{schedule_import.message}；未匹配sheet：{preview}{suffix}"
    db.flush()
    rebuild_import_item_flags(db, schedule_import)
    rebuild_import_summary(db, schedule_import)
    db.commit()
    db.refresh(schedule_import)
    return schedule_import


def _serialize_conflict_items(db: Session, item: ScheduleImportItem) -> list[dict[str, Any]]:
    if item.is_empty or item.teacher_id is None:
        return []
    rows = (
        db.query(ScheduleImportItem)
        .filter(
            ScheduleImportItem.import_id == item.import_id,
            ScheduleImportItem.id != item.id,
            ScheduleImportItem.teacher_id == item.teacher_id,
            ScheduleImportItem.weekday == item.weekday,
            ScheduleImportItem.period_id == item.period_id,
            ScheduleImportItem.is_empty == False,  # noqa: E712
        )
        .all()
    )
    conflict_items: list[dict[str, Any]] = []
    for row in rows:
        class_obj = db.query(Class).filter(Class.id == row.class_id).first()
        period = db.query(SchedulePeriod).filter(SchedulePeriod.id == row.period_id).first()
        subject = db.query(Subject).filter(Subject.id == row.subject_id).first() if row.subject_id else None
        conflict_items.append(
            {
                "id": row.id,
                "class_id": row.class_id,
                "class_name": class_obj.name if class_obj else None,
                "weekday": row.weekday,
                "period_id": row.period_id,
                "period_name": period.name if period else None,
                "subject_id": row.subject_id,
                "subject_name": subject.name if subject else None,
                "recognized_subject_name": row.recognized_subject_name,
            }
        )
    return conflict_items


def serialize_import_item(db: Session, item: ScheduleImportItem) -> dict[str, Any]:
    class_obj = db.query(Class).filter(Class.id == item.class_id).first()
    period = db.query(SchedulePeriod).filter(SchedulePeriod.id == item.period_id).first()
    subject = db.query(Subject).filter(Subject.id == item.subject_id).first() if item.subject_id else None
    teacher = db.query(User).filter(User.id == item.teacher_id).first() if item.teacher_id else None
    return {
        "id": item.id,
        "class_id": item.class_id,
        "class_name": class_obj.name if class_obj else None,
        "weekday": item.weekday,
        "period_id": item.period_id,
        "period_name": period.name if period else None,
        "subject_id": item.subject_id,
        "subject_name": subject.name if subject else None,
        "recognized_subject_name": item.recognized_subject_name,
        "teacher_id": item.teacher_id,
        "teacher_name": teacher.username if teacher else None,
        "teacher_match_status": item.teacher_match_status,
        "teacher_match_source": item.teacher_match_source,
        "teacher_candidates": _json_loads(item.teacher_candidates, []),
        "confidence": item.confidence,
        "issue_flags": _json_loads(item.issue_flags, []),
        "conflict_items": _serialize_conflict_items(db, item),
        "is_empty": item.is_empty,
    }


def update_import_item(
    db: Session,
    item: ScheduleImportItem,
    *,
    subject_id: int | None,
    teacher_id: int | None,
    is_empty: bool | None,
) -> ScheduleImportItem:
    schedule_import = db.query(ScheduleImport).filter(ScheduleImport.id == item.import_id).first()
    if is_empty is not None:
        item.is_empty = is_empty
    if subject_id is not None:
        item.subject_id = subject_id
    if teacher_id is not None:
        item.teacher_id = teacher_id
        item.teacher_match_status = "matched"
        item.teacher_match_source = "manual"
        item.teacher_candidates = _json_dumps([])
    if item.subject_id is not None and item.teacher_id is not None and not item.is_empty:
        if schedule_import:
            binding = (
                db.query(TeacherClassSubject)
                .filter(
                    TeacherClassSubject.school_id == schedule_import.school_id,
                    TeacherClassSubject.class_id == item.class_id,
                    TeacherClassSubject.subject_id == item.subject_id,
                )
                .first()
            )
            if binding is not None:
                binding.teacher_id = item.teacher_id
            else:
                db.add(
                    TeacherClassSubject(
                        school_id=schedule_import.school_id,
                        class_id=item.class_id,
                        subject_id=item.subject_id,
                        teacher_id=item.teacher_id,
                    )
                )
    item.updated_at = datetime.now()
    db.flush()
    if schedule_import:
        rebuild_import_item_flags(db, schedule_import)
        rebuild_import_summary(db, schedule_import)
    else:
        _apply_item_flags(item)
    db.commit()
    db.refresh(item)
    return item


def _raise_import_draft_conflicts(db: Session, rows: list[ScheduleImportItem]) -> None:
    teacher_slots: dict[tuple[int, int, int], ScheduleImportItem] = {}
    class_slots: dict[tuple[int, int, int], ScheduleImportItem] = {}
    teacher_conflicts: list[tuple[ScheduleImportItem, ScheduleImportItem]] = []
    class_conflicts: list[tuple[ScheduleImportItem, ScheduleImportItem]] = []

    for row in rows:
        if row.is_empty:
            continue
        teacher_key = (row.teacher_id, row.weekday, row.period_id)
        class_key = (row.class_id, row.weekday, row.period_id)
        if teacher_key in teacher_slots:
            teacher_conflicts.append((teacher_slots[teacher_key], row))
        else:
            teacher_slots[teacher_key] = row
        if class_key in class_slots:
            class_conflicts.append((class_slots[class_key], row))
        else:
            class_slots[class_key] = row

    if not teacher_conflicts and not class_conflicts:
        return

    first_conflict = teacher_conflicts[0] if teacher_conflicts else class_conflicts[0]
    first_row = first_conflict[0]
    period = db.query(SchedulePeriod).filter(SchedulePeriod.id == first_row.period_id).first()
    teacher = db.query(User).filter(User.id == first_row.teacher_id).first()
    detail_parts: list[str] = []
    if teacher_conflicts:
        detail_parts.append(
            f"教师同一时间多班冲突{len(teacher_conflicts)}处，例如 {teacher.username if teacher else '教师'} 周{first_row.weekday} {period.name if period else first_row.period_id}"
        )
    if class_conflicts:
        detail_parts.append(f"班级同一时间重复课位{len(class_conflicts)}处")
    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="；".join(detail_parts) + "。请修正后再生成待确认草案")


def create_draft_from_import(db: Session, schedule_import: ScheduleImport, *, created_by: int | None) -> ScheduleDraft:
    rows = rebuild_import_item_flags(db, schedule_import)
    blocking: list[dict[str, Any]] = []
    for row in rows:
        flags = [flag for flag in _json_loads(row.issue_flags, []) if flag != "teacher_time_conflict"]
        if flags:
            blocking.append({"code": "import_item_incomplete", "message": "导入课位仍有未处理项", "blocking": True, "entity": {"item_id": row.id, "flags": flags}})
    if blocking:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="import has unresolved items")
    _raise_import_draft_conflicts(db, rows)

    draft = ScheduleDraft(
        school_id=schedule_import.school_id,
        grade=schedule_import.grade,
        status="draft",
        score=100,
        summary=_json_dumps({"hard_pass_rate": 100, "score": 100, "locked_hits": 0, "locked_total": 0, "risk_count": 0}),
        diagnostics=_json_dumps([{"code": "schedule_import", "message": "上传课表识别生成", "blocking": False, "entity": {"import_id": schedule_import.id}}]),
        created_by=created_by,
    )
    db.add(draft)
    db.flush()
    for row in rows:
        if row.is_empty:
            continue
        db.add(
            ScheduleDraftItem(
                draft_id=draft.id,
                class_id=row.class_id,
                teacher_id=row.teacher_id,
                subject_id=row.subject_id,
                weekday=row.weekday,
                period_id=row.period_id,
                is_locked=False,
                penalty_tags=_json_dumps([]),
            )
        )
    schedule_import.status = "draft_created"
    schedule_import.message = "已生成待确认草案，正式课表尚未变更"
    db.commit()
    db.refresh(draft)
    return draft
